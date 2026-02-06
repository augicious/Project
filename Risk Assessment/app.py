from __future__ import annotations

import os
import sqlite3
import csv
import io
import uuid
import smtplib
import secrets
import time
from datetime import datetime
from functools import wraps
from pathlib import Path
from urllib.parse import urlencode
from email.message import EmailMessage

from flask import Flask, abort, redirect, render_template, request, session, url_for, flash, send_file, send_from_directory # type: ignore
from openpyxl import load_workbook
import requests
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename

try:
    from authlib.integrations.flask_client import OAuth
except Exception:  # pragma: no cover
    OAuth = None  # type: ignore

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "risks.db"
DEFAULT_RISK_REGISTER = BASE_DIR / "Risk Register.xlsx"
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", str(DATA_DIR / "uploads")))
ARCHIVE_DIR = Path(os.getenv("ATTACHMENT_ARCHIVE_DIR", str(DATA_DIR / "uploads_archive")))

DEFAULT_MAX_UPLOAD_MB = 15
try:
    _max_mb = int(float(os.getenv("UPLOAD_MAX_MB", str(DEFAULT_MAX_UPLOAD_MB))))
except ValueError:
    _max_mb = DEFAULT_MAX_UPLOAD_MB

SEVERITY_OPTIONS = ["Low", "Medium", "High", "Critical"]

TASK_STATUS_OPTIONS = ["Open", "In Progress", "Done", "Blocked", "Cancelled"]

DEFAULT_ARCHIVE_AFTER_DAYS = 180
try:
    ARCHIVE_AFTER_DAYS = int(float(os.getenv("ATTACHMENT_ARCHIVE_AFTER_DAYS", str(DEFAULT_ARCHIVE_AFTER_DAYS))))
except ValueError:
    ARCHIVE_AFTER_DAYS = DEFAULT_ARCHIVE_AFTER_DAYS


def _env_bool(name: str, default: bool = False) -> bool:
    val = os.getenv(name)
    if val is None:
        return default
    return str(val).strip().lower() in {"1", "true", "yes", "on"}


def _parse_recipients(value: str | None) -> list[str]:
    if not value:
        return []
    parts = [p.strip() for p in value.replace(";", ",").split(",")]
    return [p for p in parts if p]


def _normalize_email(value: str | None) -> str:
    v = (value or "").strip().lower()
    if not v:
        return ""
    # Accept common forms like:
    #   - user@domain
    #   - Display Name <user@domain>
    #   - user@domain; other@domain
    # Extract the first email-looking substring.
    import re

    m = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", v, flags=re.IGNORECASE)
    if not m:
        return ""
    return m.group(0).strip().lower()


def _send_email(subject: str, html_body: str, to_addrs: list[str]) -> bool:
    """Send an email using SMTP settings from env vars.

    Required env vars:
      - SMTP_HOST
      - SMTP_PORT (default 587)
      - SMTP_USER (optional)
      - SMTP_PASSWORD (optional)
      - SMTP_USE_TLS (default true)
      - SMTP_FROM (default noreply@hdh.org)
      - SMTP_REPLY_TO (default VFSA@hdh.org)

    If SMTP_HOST is missing or to_addrs is empty, this is a no-op and returns False.
    """

    smtp_host = os.getenv("SMTP_HOST", "").strip()
    if not smtp_host:
        return False
    if not to_addrs:
        return False

    smtp_port = int(float(os.getenv("SMTP_PORT", "587")))

    # Backward-compatible env var names
    smtp_user = (os.getenv("SMTP_USER") or os.getenv("SMTP_USERNAME") or os.getenv("SMTP_LOGIN") or "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD") or os.getenv("SMTP_PASS") or os.getenv("SMTP_PWD") or ""
    smtp_use_tls = _env_bool("SMTP_USE_TLS", _env_bool("SMTP_STARTTLS", True))

    mail_from = (os.getenv("SMTP_FROM") or os.getenv("MAIL_FROM") or "noreply@hdh.org").strip() or "noreply@hdh.org"
    reply_to = (os.getenv("SMTP_REPLY_TO") or os.getenv("REPLY_TO") or "VFSA@hdh.org").strip()

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = mail_from
    msg["To"] = ", ".join(to_addrs)
    if reply_to:
        msg["Reply-To"] = reply_to
    msg.set_content("This message requires an HTML capable email client.")
    msg.add_alternative(html_body, subtype="html")

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            if smtp_use_tls:
                server.starttls()
            if smtp_user:
                server.login(smtp_user, smtp_password)
            server.send_message(msg)
        return True
    except Exception:
        return False


def _unique_emails(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for v in values:
        addr = _normalize_email(v)
        if not addr:
            continue
        if addr in seen:
            continue
        seen.add(addr)
        out.append(addr)
    return out


def _current_actor() -> str:
    user = session.get("user") or {}
    actor = (user.get("preferred_username") or user.get("name") or "").strip()
    if actor:
        return actor
    if session.get("is_admin"):
        return "admin"
    return "system"


def _current_user_email() -> str:
    """Best-effort current user email for "assigned=me" style filters.

    When OIDC is enabled, this comes from the OIDC session user.
    When admin-password mode is used without OIDC/AUTH_REQUIRED, allow a stored
    email (set during admin login) so "My queue" can still function.
    """

    user = session.get("user") or {}
    candidates = [
        (user or {}).get("preferred_username"),
        (user or {}).get("email"),
        session.get("admin_user_email"),
        # Common reverse-proxy identity headers (optional)
        request.headers.get("X-MS-CLIENT-PRINCIPAL-NAME"),
        request.headers.get("X-Forwarded-User"),
        request.headers.get("X-Remote-User"),
        request.environ.get("REMOTE_USER"),
    ]
    for c in candidates:
        email = _normalize_email(str(c) if c is not None else "")
        if email:
            return email
    return ""


def _assignee_me_condition(current_user_email: str) -> tuple[str, list[str]]:
    """Return a SQL condition + params that match "assigned_to" for the current user.

    This is intentionally tolerant:
    - Matches exact values (case-insensitive) for name / local-part.
    - Also matches email embedded in a larger string (e.g. "Name <email>").
    """

    clauses: list[str] = []
    params: list[str] = []

    email = _normalize_email(current_user_email)
    if email:
        clauses.append("LOWER(COALESCE(assigned_to,'')) LIKE ?")
        params.append(f"%{email}%")

        local = (email.split("@", 1)[0] or "").strip().lower()
        if local:
            clauses.append("LOWER(TRIM(COALESCE(assigned_to,''))) = ?")
            params.append(local)

    display_name = str((session.get("user") or {}).get("name") or "").strip().lower()
    if display_name:
        clauses.append("LOWER(TRIM(COALESCE(assigned_to,''))) = ?")
        params.append(display_name)

    if not clauses:
        return "", []
    return "(" + " OR ".join(clauses) + ")", params


def _load_dotenv_if_available() -> None:
    """Load environment variables from a .env file if python-dotenv is installed.

    - If ENV_FILE is set, loads that file.
    - Otherwise, loads a .env from the current working directory (service sets CWD to project root).
    """

    try:
        from dotenv import load_dotenv  # type: ignore
    except Exception:
        return

    env_file = os.getenv("ENV_FILE", "").strip()
    if env_file:
        # For services, ENV_FILE should be the source of truth.
        # Use override=True so blank/stale machine env vars don't win.
        load_dotenv(dotenv_path=env_file, override=True)
    else:
        load_dotenv(override=False)


_load_dotenv_if_available()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("FLASK_SECRET_KEY", "dev-secret-change-me")
app.config["ADMIN_PASSWORD"] = os.getenv("ADMIN_PASSWORD", "admin")
app.config["APP_VERSION"] = "2026-01-19-01"
app.config["MAX_CONTENT_LENGTH"] = max(1, _max_mb) * 1024 * 1024
app.config["ATTACHMENT_ARCHIVE_AFTER_DAYS"] = ARCHIVE_AFTER_DAYS
app.config["RISK_REGISTER_PATH"] = Path(
    os.getenv("RISK_REGISTER_PATH", str(DEFAULT_RISK_REGISTER))
)
app.config["DB_INITIALIZED"] = False
app.config["ENABLE_ADMIN_IMPORT"] = _env_bool("ENABLE_ADMIN_IMPORT", False)

# Trust reverse-proxy headers (IIS/ARR) so redirects and OAuth callbacks work behind HTTPS.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)


@app.get("/favicon.ico")
def favicon():
    icons_dir = os.path.join(app.root_path, "static", "icons")
    return send_from_directory(
        icons_dir,
        "favicon.ico",
        mimetype="image/vnd.microsoft.icon",
        max_age=60 * 60 * 24 * 7,
    )

# Microsoft Entra ID (Microsoft 365) OpenID Connect settings (configure via env vars)
app.config["OIDC_TENANT_ID"] = os.getenv("OIDC_TENANT_ID", "").strip()
app.config["OIDC_CLIENT_ID"] = os.getenv("OIDC_CLIENT_ID", "").strip()
app.config["OIDC_CLIENT_SECRET"] = os.getenv("OIDC_CLIENT_SECRET", "").strip()
app.config["OIDC_REDIRECT_URI"] = os.getenv("OIDC_REDIRECT_URI", "").strip()
app.config["OIDC_POST_LOGOUT_REDIRECT_URI"] = os.getenv(
    "OIDC_POST_LOGOUT_REDIRECT_URI", ""
).strip()

app.config["AUTH_REQUIRED"] = os.getenv("AUTH_REQUIRED", "true").lower() in {
    "1",
    "true",
    "yes",
    "on",
}

# Session hardening (safe defaults; override via env if needed)
_debug_mode = os.getenv("FLASK_DEBUG", "").strip().lower() in {"1", "true", "yes", "on"}
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = os.getenv("SESSION_COOKIE_SAMESITE", "Lax")
app.config["SESSION_COOKIE_SECURE"] = _env_bool("SESSION_COOKIE_SECURE", not _debug_mode)

# Admin authorization via Entra group membership
app.config["ADMIN_ENTRA_GROUPS"] = os.getenv(
    "ADMIN_ENTRA_GROUPS",
    "Risk Management,IS-Administrator",
)
app.config["ADMIN_ENTRA_GROUP_IDS"] = os.getenv("ADMIN_ENTRA_GROUP_IDS", "").strip()
app.config["ADMIN_ENTRA_GRAPH_LOOKUP"] = _env_bool("ADMIN_ENTRA_GRAPH_LOOKUP", True)


def _oauth_enabled() -> bool:
    return (
        OAuth is not None
        and bool(app.config.get("OIDC_TENANT_ID"))
        and bool(app.config.get("OIDC_CLIENT_ID"))
        and bool(app.config.get("OIDC_CLIENT_SECRET"))
        and bool(app.config.get("OIDC_REDIRECT_URI"))
    )


# Admin password login mode: keep for non-OIDC deployments, but default to disabled when OIDC is enabled.
app.config["ADMIN_PASSWORD_ENABLED"] = _env_bool("ADMIN_PASSWORD_ENABLED", not _oauth_enabled())


oauth = OAuth(app) if OAuth is not None else None
if oauth is not None and _oauth_enabled():
    tenant = app.config["OIDC_TENANT_ID"]
    metadata_url = (
        f"https://login.microsoftonline.com/{tenant}/v2.0/.well-known/openid-configuration"
    )
    oauth.register(
        name="microsoft",
        server_metadata_url=metadata_url,
        client_id=app.config["OIDC_CLIENT_ID"],
        client_secret=app.config["OIDC_CLIENT_SECRET"],
        client_kwargs={
            # Customize via env var OIDC_SCOPES.
            # Recommended if using Graph-based group lookup: add GroupMember.Read.All (requires admin consent).
            "scope": os.getenv("OIDC_SCOPES", "openid profile email User.Read"),
        },
    )


def _split_csv(value: str) -> list[str]:
    parts = [p.strip() for p in (value or "").split(",")]
    return [p for p in parts if p]


def _csrf_token() -> str:
    token = session.get("_csrf_token")
    if token and isinstance(token, str):
        return token
    token = secrets.token_urlsafe(32)
    session["_csrf_token"] = token
    return token


@app.context_processor
def _inject_template_flags():
    return {
        "enable_admin_import": bool(app.config.get("ENABLE_ADMIN_IMPORT", False)),
    }


app.jinja_env.globals["csrf_token"] = _csrf_token


@app.before_request
def _csrf_protect():
    # Enforce CSRF on state-changing requests.
    if request.method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return None
    if request.path.startswith("/static/"):
        return None

    sent = (
        request.form.get("_csrf_token")
        or request.headers.get("X-CSRF-Token")
        or request.headers.get("X-CSRFToken")
    )
    expected = session.get("_csrf_token")
    if not sent or not expected or str(sent) != str(expected):
        abort(400)
    return None


@app.after_request
def _security_headers(response):
    # Basic defense-in-depth headers.
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault(
        "Permissions-Policy",
        "camera=(), microphone=(), geolocation=(), interest-cohort=()",
    )

    if _env_bool("CSP_ENABLED", True):
        # Keep permissive enough for inline scripts already used in templates.
        response.headers.setdefault(
            "Content-Security-Policy",
            " ".join(
                [
                    "default-src 'self';",
                    "base-uri 'self';",
                    "frame-ancestors 'self';",
                    "img-src 'self' data:;",
                    "style-src 'self' 'unsafe-inline';",
                    "script-src 'self' 'unsafe-inline';",
                    "object-src 'none';",
                ]
            ),
        )
    return response


def _graph_get_member_of_group_names(access_token: str) -> list[str]:
    if not access_token:
        return []

    url = "https://graph.microsoft.com/v1.0/me/memberOf?$select=displayName&$top=999"
    headers = {"Authorization": f"Bearer {access_token}"}
    names: list[str] = []

    for _ in range(20):
        resp = requests.get(url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return []
        data = resp.json() or {}
        for item in (data.get("value") or []):
            dn = (item or {}).get("displayName")
            if dn:
                names.append(str(dn))
        next_link = data.get("@odata.nextLink")
        if not next_link:
            break
        url = str(next_link)

    # De-dup preserving order
    seen: set[str] = set()
    out: list[str] = []
    for n in names:
        key = n.strip().lower()
        if key and key not in seen:
            seen.add(key)
            out.append(n)
    return out


_ASSIGNEE_CACHE: dict[str, object] = {
    "expires": 0.0,
    "emails": [],
    "source": "",
}


def _assignee_cache_db_read() -> tuple[list[str], str, float]:
    """Read allowlist cache from SQLite for multi-process stability."""

    try:
        with get_db_connection() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS app_cache (
                    cache_key TEXT PRIMARY KEY,
                    cache_value TEXT NOT NULL,
                    expires_epoch REAL NOT NULL,
                    updated_epoch REAL NOT NULL
                )
                """
            )
            row = conn.execute(
                "SELECT cache_value, expires_epoch FROM app_cache WHERE cache_key = ?",
                ("assignee_allowlist",),
            ).fetchone()
            if not row:
                return [], "", 0.0

            cache_value = str(row["cache_value"] or "")
            try:
                expires = float(row["expires_epoch"] or 0.0)
            except Exception:
                expires = 0.0

            import json

            payload = json.loads(cache_value) if cache_value else {}
            if not isinstance(payload, dict):
                return [], "", expires
            emails = payload.get("emails")
            source = str(payload.get("source") or "")
            if not isinstance(emails, list):
                return [], source, expires
            return [str(x) for x in emails], source, expires
    except Exception:
        return [], "", 0.0


def _assignee_cache_db_write(*, emails: list[str], source: str, expires_epoch: float) -> None:
    """Persist allowlist cache to SQLite. Best-effort."""

    if not emails:
        return
    try:
        import json

        payload = json.dumps({"emails": emails, "source": source}, separators=(",", ":"))
        now = time.time()
        with get_db_connection() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS app_cache (
                    cache_key TEXT PRIMARY KEY,
                    cache_value TEXT NOT NULL,
                    expires_epoch REAL NOT NULL,
                    updated_epoch REAL NOT NULL
                )
                """
            )
            conn.execute(
                """
                INSERT INTO app_cache(cache_key, cache_value, expires_epoch, updated_epoch)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(cache_key) DO UPDATE SET
                    cache_value=excluded.cache_value,
                    expires_epoch=excluded.expires_epoch,
                    updated_epoch=excluded.updated_epoch
                """,
                ("assignee_allowlist", payload, float(expires_epoch), float(now)),
            )
            conn.commit()
    except Exception:
        return


def _graph_app_token() -> str:
    """Get an app-only Microsoft Graph token (client credentials).

    Requires:
      - OIDC_TENANT_ID
      - OIDC_CLIENT_ID
      - OIDC_CLIENT_SECRET

    And Graph application permissions with admin consent to enumerate group members.
    """

    tenant = str(app.config.get("OIDC_TENANT_ID") or "").strip()
    client_id = str(app.config.get("OIDC_CLIENT_ID") or "").strip()
    client_secret = str(app.config.get("OIDC_CLIENT_SECRET") or "").strip()
    if not (tenant and client_id and client_secret):
        return ""

    token_url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
    data = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "client_credentials",
        "scope": "https://graph.microsoft.com/.default",
    }

    try:
        resp = requests.post(token_url, data=data, timeout=10)
        if resp.status_code != 200:
            try:
                logging.warning("Graph app token request failed (status=%s)", resp.status_code)
            except Exception:
                pass
            return ""
        payload = resp.json() or {}
        return str(payload.get("access_token") or "").strip()
    except Exception:
        try:
            logging.exception("Graph app token request raised exception")
        except Exception:
            pass
        return ""


def _graph_group_ids_by_names(*, access_token: str, display_names: list[str]) -> list[str]:
    if not access_token:
        return []
    names = [str(n or "").strip() for n in display_names if str(n or "").strip()]
    if not names:
        return []

    headers = {"Authorization": f"Bearer {access_token}"}
    out: list[str] = []
    seen: set[str] = set()
    for name in names:
        # Filter requires single quotes; escape per OData rules.
        safe = name.replace("'", "''")
        try:
            resp = requests.get(
                "https://graph.microsoft.com/v1.0/groups",
                headers=headers,
                params={
                    "$select": "id,displayName",
                    "$filter": f"displayName eq '{safe}'",
                    "$top": "50",
                },
                timeout=10,
            )
            if resp.status_code != 200:
                try:
                    logging.warning(
                        "Graph group lookup failed (status=%s, name=%s)",
                        resp.status_code,
                        name,
                    )
                except Exception:
                    pass
                continue
            data = resp.json() or {}
            for item in (data.get("value") or []):
                gid = str((item or {}).get("id") or "").strip()
                if gid and gid.lower() not in seen:
                    seen.add(gid.lower())
                    out.append(gid)
        except Exception:
            continue
    return out


def _graph_group_member_emails(*, access_token: str, group_id: str) -> list[str]:
    if not access_token or not group_id:
        return []

    headers = {"Authorization": f"Bearer {access_token}"}
    base = "https://graph.microsoft.com/v1.0/groups/" + str(group_id).strip()
    # Prefer transitiveMembers (includes nested groups). Some tenants restrict this; fall back to /members.
    url = base + "/transitiveMembers/microsoft.graph.user" + "?$select=mail,userPrincipalName,displayName&$top=999"
    emails: list[str] = []

    tried_fallback = False

    for _ in range(25):
        try:
            resp = requests.get(url, headers=headers, timeout=15)
            if resp.status_code != 200:
                try:
                    logging.warning(
                        "Graph group members fetch failed (status=%s, group_id=%s, url=%s)",
                        resp.status_code,
                        str(group_id)[:8] + "...",
                        ("transitive" if "/transitiveMembers/" in url else "members"),
                    )
                except Exception:
                    pass
                if (not tried_fallback) and "/transitiveMembers/" in url:
                    tried_fallback = True
                    url = base + "/members/microsoft.graph.user" + "?$select=mail,userPrincipalName,displayName&$top=999"
                    continue
                break
            data = resp.json() or {}
            for item in (data.get("value") or []):
                mail = (item or {}).get("mail")
                upn = (item or {}).get("userPrincipalName")
                addr = _normalize_email(str(mail or upn or ""))
                if addr:
                    emails.append(addr)
            next_link = data.get("@odata.nextLink")
            if not next_link:
                break
            url = str(next_link)
        except Exception:
            break

    return _unique_emails(emails)


def _allowed_assignee_emails(*, force_refresh: bool = False) -> tuple[list[str], str]:
    """Assignee allowlist (normalized emails).

    Sources (first that produces entries wins):
      1) ASSIGNEE_ALLOWLIST (CSV/semicolon)
      2) Entra group members via Graph app-only token:
         - ASSIGNEE_ENTRA_GROUP_IDS (CSV of GUIDs)
         - or ASSIGNEE_ENTRA_GROUPS (CSV of display names; default: Risk Management)
    """

    try:
        ttl = int(float(os.getenv("ASSIGNEE_CACHE_SECONDS", "300")))
    except ValueError:
        ttl = 300
    ttl = max(30, min(ttl, 60 * 60))
    try:
        stale_ttl = int(float(os.getenv("ASSIGNEE_STALE_SECONDS", str(ttl))))
    except ValueError:
        stale_ttl = ttl
    stale_ttl = max(30, min(stale_ttl, 60 * 60))

    now = time.time()

    explicit_group_ids_raw = os.getenv("ASSIGNEE_ENTRA_GROUP_IDS", "").strip()
    explicit_group_names_raw = os.getenv("ASSIGNEE_ENTRA_GROUPS", "").strip()
    explicit_group_configured = bool(explicit_group_ids_raw or explicit_group_names_raw)

    # Read env allowlist only when Entra group is NOT explicitly configured.
    # If a group is configured, ASSIGNEE_ALLOWLIST should never override or act as fallback.
    env_list = os.getenv("ASSIGNEE_ALLOWLIST", "").strip() if (not explicit_group_configured) else ""
    env_fallback_emails = _unique_emails(_parse_recipients(env_list)) if env_list else []

    # Snapshot current cache (may be used as last-known-good).
    existing_emails: list[str] = []
    existing_source = ""
    existing_expires = 0.0
    try:
        existing_expires = float(_ASSIGNEE_CACHE.get("expires") or 0.0)
    except Exception:
        existing_expires = 0.0
    cached = _ASSIGNEE_CACHE.get("emails") or []
    if isinstance(cached, list):
        existing_emails = [str(x) for x in cached if str(x).strip()]
    existing_source = str(_ASSIGNEE_CACHE.get("source") or "")

    if (not force_refresh) and existing_expires and now < existing_expires and existing_emails:
        if explicit_group_configured and str(existing_source).startswith("env:ASSIGNEE_ALLOWLIST"):
            # Ignore env-derived cache when Entra group config is explicit.
            pass
        else:
            return existing_emails, existing_source

    # If in-memory cache is missing/expired, try DB-backed cache (helps with multi-process deployments).
    if (not force_refresh) and (not existing_emails or not existing_expires or now >= existing_expires):
        db_emails, db_source, db_expires = _assignee_cache_db_read()
        if db_emails and db_expires and now < db_expires:
            if explicit_group_configured and str(db_source).startswith("env:ASSIGNEE_ALLOWLIST"):
                # Ignore env-derived DB cache when Entra group config is explicit.
                pass
            else:
                _ASSIGNEE_CACHE["expires"] = db_expires
                _ASSIGNEE_CACHE["emails"] = db_emails
                _ASSIGNEE_CACHE["source"] = db_source
                return db_emails, db_source

    if env_fallback_emails and (not explicit_group_configured):
        expires_epoch = now + ttl
        _ASSIGNEE_CACHE["expires"] = expires_epoch
        _ASSIGNEE_CACHE["emails"] = env_fallback_emails
        _ASSIGNEE_CACHE["source"] = "env:ASSIGNEE_ALLOWLIST"
        _assignee_cache_db_write(emails=env_fallback_emails, source="env:ASSIGNEE_ALLOWLIST", expires_epoch=expires_epoch)
        return env_fallback_emails, "env:ASSIGNEE_ALLOWLIST"

    group_ids = _split_csv(explicit_group_ids_raw)
    group_names = _split_csv(explicit_group_names_raw)
    if not group_ids and not group_names:
        group_names = ["Risk Management"]

    emails: list[str] = []
    if group_ids or group_names:
        access_token = _graph_app_token()
        resolved_ids: list[str] = []
        resolved_ids.extend([g for g in group_ids if g])
        if access_token and group_names:
            resolved_ids.extend(_graph_group_ids_by_names(access_token=access_token, display_names=group_names))
        if access_token and resolved_ids:
            for gid in resolved_ids:
                emails.extend(_graph_group_member_emails(access_token=access_token, group_id=gid))
        emails = _unique_emails(emails)

    if emails:
        expires_epoch = now + ttl
        _ASSIGNEE_CACHE["expires"] = expires_epoch
        _ASSIGNEE_CACHE["emails"] = emails
        _ASSIGNEE_CACHE["source"] = "entra"
        _assignee_cache_db_write(emails=emails, source="entra", expires_epoch=expires_epoch)
        return emails, "entra"

    # If Entra lookup is configured and returns empty, do not fall back to ASSIGNEE_ALLOWLIST.

    # Graph/enumeration returned empty. Do not wipe out a previously-good cache entry.
    if existing_emails:
        expires_epoch = now + stale_ttl
        stale_source = (existing_source or "cached")
        if "stale" not in stale_source.lower():
            stale_source = f"{stale_source} (stale)"
        _ASSIGNEE_CACHE["expires"] = expires_epoch
        _ASSIGNEE_CACHE["emails"] = existing_emails
        _ASSIGNEE_CACHE["source"] = stale_source
        _assignee_cache_db_write(emails=existing_emails, source=stale_source, expires_epoch=expires_epoch)
        return existing_emails, stale_source

    _ASSIGNEE_CACHE["expires"] = now + ttl
    _ASSIGNEE_CACHE["emails"] = []
    _ASSIGNEE_CACHE["source"] = "entra(empty)"
    return [], "entra(empty)"


def _validate_assignee_email(value: str) -> tuple[bool, str]:
    """Validate & normalize assigned_to.

    - Returns (ok, normalized_email_or_blank)
    - Enforces allowlist when it is available (non-empty)
    """

    raw = (value or "").strip()
    if not raw:
        return True, ""

    email = _normalize_email(raw)
    if not email:
        return False, ""

    allowlist, _src = _allowed_assignee_emails()
    if allowlist:
        allowed = {a.strip().lower() for a in allowlist}
        if email.strip().lower() not in allowed:
            return False, email

    return True, email


def _compute_is_admin_from_entra(*, token: dict, userinfo: dict) -> bool:
    allowed_names = {n.strip().lower() for n in _split_csv(app.config.get("ADMIN_ENTRA_GROUPS", ""))}
    allowed_ids = {n.strip().lower() for n in _split_csv(app.config.get("ADMIN_ENTRA_GROUP_IDS", ""))}

    if not allowed_names and not allowed_ids:
        return False

    # If the ID token contains a 'groups' claim, it is usually a list of GUIDs.
    groups_claim = userinfo.get("groups")
    if isinstance(groups_claim, list) and allowed_ids:
        for gid in groups_claim:
            if str(gid).strip().lower() in allowed_ids:
                return True

    if not app.config.get("ADMIN_ENTRA_GRAPH_LOOKUP", True):
        return False

    access_token = str(token.get("access_token") or "").strip()
    group_names = _graph_get_member_of_group_names(access_token)
    for name in group_names:
        if name.strip().lower() in allowed_names:
            return True
    return False


def is_authenticated() -> bool:
    return bool(session.get("user"))


def is_admin() -> bool:
    """Return True when the current session has admin rights."""

    if session.get("is_admin"):
        return True

    user = session.get("user")
    if isinstance(user, dict) and user.get("is_admin"):
        return True

    return False


def require_login() -> bool:
    if not _oauth_enabled() or not app.config.get("AUTH_REQUIRED", True):
        return True
    if is_authenticated():
        return True
    return False


def require_admin() -> bool:
    """Return True when the request is allowed to access admin routes."""

    # When admin-password mode is enabled, a local session flag grants access.
    if app.config.get("ADMIN_PASSWORD_ENABLED", True) and session.get("is_admin"):
        return True

    # Otherwise, rely on Entra group-based admin (stored on the OIDC session user).
    if _oauth_enabled() and app.config.get("AUTH_REQUIRED", True):
        return is_admin()

    # No auth configured: fall back to session flag only.
    return bool(session.get("is_admin"))


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if require_login():
            return view_func(*args, **kwargs)
        return redirect(url_for("login", next=request.full_path))

    return wrapper


@app.before_request
def _enforce_site_login():
    # Only enforce when OIDC is configured and AUTH_REQUIRED is enabled.
    if not _oauth_enabled() or not app.config.get("AUTH_REQUIRED", True):
        return None

    # Allow health checks and auth endpoints
    if request.path.startswith("/static/"):
        return None
    if request.endpoint in {
        "health",
        "login",
        "auth_callback",
        "logout",
    }:
        return None

    if is_authenticated():
        return None

    return redirect(url_for("login", next=request.full_path))


@app.errorhandler(413)
def request_entity_too_large(_err):
    flash(
        f"Upload too large. Limit is {app.config.get('MAX_CONTENT_LENGTH', 0) // (1024 * 1024)}MB per request.",
        "warning",
    )
    return redirect(request.referrer or url_for("admin_dashboard"))


@app.errorhandler(400)
def bad_request(_err):
    # Most common 400 in this app will be missing/invalid CSRF.
    flash("Bad request. Please refresh the page and try again.", "warning")
    return redirect(request.referrer or url_for("index"))


@app.route("/login")
def login():
    if not _oauth_enabled():
        flash(
            "Microsoft 365 authentication is not configured. Set OIDC_TENANT_ID, OIDC_CLIENT_ID, OIDC_CLIENT_SECRET, and OIDC_REDIRECT_URI.",
            "warning",
        )
        return redirect(url_for("index"))

    if is_authenticated():
        return redirect(request.args.get("next") or url_for("index"))

    # Save where to return after sign-in (relative URLs only).
    next_url = request.args.get("next")
    if next_url and next_url.startswith("/"):
        session["post_login_redirect"] = next_url

    # Authlib uses the session to store state/nonce.
    redirect_uri = app.config["OIDC_REDIRECT_URI"]
    return oauth.microsoft.authorize_redirect(redirect_uri=redirect_uri)  # type: ignore


@app.route("/auth/callback")
def auth_callback():
    if not _oauth_enabled():
        flash("Microsoft 365 authentication is not configured.", "danger")
        return redirect(url_for("index"))

    try:
        token = oauth.microsoft.authorize_access_token()  # type: ignore
    except Exception:
        flash("Sign-in failed. Please try again.", "danger")
        return redirect(url_for("index"))

    # Clear any stale/large session payload from older versions or intermediary auth state.
    # Keep the intended post-login redirect.
    next_url = session.pop("post_login_redirect", "")
    session.clear()
    if next_url:
        session["post_login_redirect"] = next_url

    userinfo = token.get("userinfo")
    if not userinfo:
        try:
            userinfo = oauth.microsoft.parse_id_token(token)  # type: ignore
        except Exception:
            userinfo = {}

    session["user"] = {
        "name": userinfo.get("name") or "",
        "preferred_username": userinfo.get("preferred_username")
        or userinfo.get("upn")
        or userinfo.get("email")
        or "",
        "oid": userinfo.get("oid") or userinfo.get("sub") or "",
    }

    # Evaluate admin rights based on Entra group membership.
    try:
        is_entra_admin = _compute_is_admin_from_entra(token=token or {}, userinfo=userinfo or {})
    except Exception:
        is_entra_admin = False
    session["user"]["is_admin"] = bool(is_entra_admin)
    if is_entra_admin:
        session["is_admin"] = True
    else:
        session.pop("is_admin", None)

    next_url = session.pop("post_login_redirect", "")
    if next_url and next_url.startswith("/"):
        return redirect(next_url)
    return redirect(url_for("index"))


@app.route("/logout")
def logout():
    session.pop("user", None)
    session.pop("is_admin", None)
    session.pop("_csrf_token", None)

    if not _oauth_enabled():
        return redirect(url_for("index"))

    tenant = app.config["OIDC_TENANT_ID"]
    base = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/logout"

    post_logout = app.config.get("OIDC_POST_LOGOUT_REDIRECT_URI")
    if not post_logout:
        # Fall back to app root (prefer https via ProxyFix + forwarded headers)
        post_logout = request.url_root.rstrip("/")

    return redirect(f"{base}?{urlencode({'post_logout_redirect_uri': post_logout})}")


def get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    with get_db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS risks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                description TEXT NOT NULL,
                likelihood TEXT,
                impact TEXT,
                likelihood_initial TEXT,
                impact_initial TEXT,
                likelihood_residual TEXT,
                impact_residual TEXT,
                owner TEXT,
                status TEXT NOT NULL DEFAULT 'New',
                mitigation TEXT,
                progress INTEGER NOT NULL DEFAULT 0,
                admin_notes TEXT,
                date_identified TEXT,
                priority INTEGER,
                impact_type TEXT,
                review_period TEXT,
                date_last_reviewed TEXT,
                next_review TEXT,
                date_postponed TEXT,
                reason_postponed TEXT,
                date_mitigated TEXT,
                mitigation_history TEXT,
                source_sheet TEXT,
                severity TEXT,
                assigned_to TEXT,
                created_by TEXT,
                updated_by TEXT,
                closed_at TEXT,
                close_reason TEXT,
                deleted_at TEXT,
                deleted_by TEXT,
                deleted_reason TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS risk_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                risk_id INTEGER NOT NULL,
                event_type TEXT NOT NULL,
                field TEXT,
                old_value TEXT,
                new_value TEXT,
                actor TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS risk_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                risk_id INTEGER NOT NULL,
                storage_name TEXT NOT NULL,
                original_name TEXT NOT NULL,
                content_type TEXT,
                size_bytes INTEGER,
                uploaded_by TEXT,
                created_at TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS risk_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                risk_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                notes TEXT,
                status TEXT NOT NULL DEFAULT 'Open',
                assigned_to TEXT,
                due_date TEXT,
                created_by TEXT,
                updated_by TEXT,
                completed_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS risk_comments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                risk_id INTEGER NOT NULL,
                body TEXT NOT NULL,
                author TEXT,
                created_at TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS risk_watchers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                risk_id INTEGER NOT NULL,
                email TEXT NOT NULL,
                created_by TEXT,
                created_at TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS risk_score_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                snapshot_date TEXT NOT NULL,
                risk_id INTEGER NOT NULL,
                score_initial INTEGER,
                level_initial TEXT,
                score_residual INTEGER,
                level_residual TEXT,
                created_by TEXT,
                created_at TEXT NOT NULL
            )
            """
        )

        # Lightweight migrations (safe on existing DBs)
        cols = [row[1] for row in conn.execute("PRAGMA table_info(risk_attachments)").fetchall()]
        if "is_archived" not in cols:
            conn.execute("ALTER TABLE risk_attachments ADD COLUMN is_archived INTEGER NOT NULL DEFAULT 0")
        if "archived_at" not in cols:
            conn.execute("ALTER TABLE risk_attachments ADD COLUMN archived_at TEXT")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS kpis (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                value TEXT,
                notes TEXT,
                source_sheet TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        extra_columns = [
            ("date_identified", "TEXT"),
            ("priority", "INTEGER"),
            ("impact_type", "TEXT"),
            ("likelihood_initial", "TEXT"),
            ("impact_initial", "TEXT"),
            ("likelihood_residual", "TEXT"),
            ("impact_residual", "TEXT"),
            ("review_period", "TEXT"),
            ("date_last_reviewed", "TEXT"),
            ("next_review", "TEXT"),
            ("date_postponed", "TEXT"),
            ("reason_postponed", "TEXT"),
            ("date_mitigated", "TEXT"),
            ("mitigation_history", "TEXT"),
            ("source_sheet", "TEXT"),
            ("severity", "TEXT"),
            ("assigned_to", "TEXT"),
            ("created_by", "TEXT"),
            ("updated_by", "TEXT"),
            ("closed_at", "TEXT"),
            ("close_reason", "TEXT"),
            ("deleted_at", "TEXT"),
            ("deleted_by", "TEXT"),
            ("deleted_reason", "TEXT"),
        ]
        for column_name, column_type in extra_columns:
            try:
                conn.execute(
                    f"ALTER TABLE risks ADD COLUMN {column_name} {column_type}"
                )
            except sqlite3.OperationalError:
                pass

        # Indexes (safe no-ops if they already exist). These dramatically help filter/sort performance.
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risks_updated_at ON risks(updated_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risks_status ON risks(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risks_severity ON risks(severity)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risks_owner ON risks(owner)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risks_assigned_to ON risks(assigned_to)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risks_source_sheet ON risks(source_sheet)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risks_deleted_at ON risks(deleted_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risks_deleted_by ON risks(deleted_by)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risk_events_risk_type_created ON risk_events(risk_id, event_type, created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risk_attachments_risk_id ON risk_attachments(risk_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risk_tasks_risk_id ON risk_tasks(risk_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risk_tasks_status ON risk_tasks(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risk_tasks_due_date ON risk_tasks(due_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risk_tasks_assigned_to ON risk_tasks(assigned_to)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risk_comments_risk_id ON risk_comments(risk_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risk_comments_created_at ON risk_comments(created_at)")
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_risk_watchers_unique ON risk_watchers(risk_id, email)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_risk_score_snapshots_date ON risk_score_snapshots(snapshot_date)")
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_risk_score_snapshots_unique ON risk_score_snapshots(snapshot_date, risk_id)")
        conn.commit()


def _parse_pagination(*, default_per_page: int = 50, max_per_page: int = 200) -> tuple[int, int]:
    try:
        page = int(float(request.args.get("page", "1")))
    except ValueError:
        page = 1
    try:
        per_page = int(float(request.args.get("per_page", str(default_per_page))))
    except ValueError:
        per_page = default_per_page

    page = max(1, page)
    per_page = max(1, min(max_per_page, per_page))
    return page, per_page


def _pager(*, total: int, page: int, per_page: int) -> dict[str, int | bool | str | None]:
    pages = max(1, (max(0, total) + per_page - 1) // per_page)
    page = min(max(1, page), pages)
    return {
        "page": page,
        "per_page": per_page,
        "total": total,
        "pages": pages,
        "has_prev": page > 1,
        "has_next": page < pages,
        "prev_url": None,
        "next_url": None,
    }


def _where_join(base_where_sql: str, condition_sql: str) -> str:
    if not condition_sql.strip():
        return base_where_sql
    if base_where_sql.strip():
        return base_where_sql + " AND " + condition_sql
    return " WHERE " + condition_sql


def _not_deleted_condition(table_alias: str | None = None) -> str:
    col = f"{table_alias}.deleted_at" if table_alias else "deleted_at"
    return f"TRIM(COALESCE({col}, '')) = ''"


def _normalize_header(header: str) -> str:
    return "".join(ch.lower() for ch in header if ch.isalnum())


def seed_from_excel(force: bool = False) -> int:
    if not app.config["RISK_REGISTER_PATH"].exists():
        return 0

    if not force:
        with get_db_connection() as conn:
            existing = conn.execute("SELECT COUNT(*) FROM risks").fetchone()[0]
            if existing:
                return 0

    workbook = load_workbook(app.config["RISK_REGISTER_PATH"], data_only=True)

    header_aliases = {
        "owner": ["Risk Owner", "Owner", "Responsible", "Assignee"],
        "date_identified": ["Date Identified"],
        "priority": ["Priority (1-5)", "Priority"],
        "description": ["Description of Risk", "Description", "Risk Description"],
        "likelihood": ["Likelihood", "Risk Likelihood"],
        "impact": ["Impact", "Risk Impact"],
        "likelihood_initial": [
            "Likelihood (Initial)",
            "Initial Likelihood",
            "Inherent Likelihood",
            "Pre-Mitigation Likelihood",
        ],
        "impact_initial": [
            "Impact (Initial)",
            "Initial Impact",
            "Inherent Impact",
            "Pre-Mitigation Impact",
        ],
        "likelihood_residual": [
            "Likelihood (Residual)",
            "Residual Likelihood",
            "Post-Mitigation Likelihood",
            "Current Likelihood",
        ],
        "impact_residual": [
            "Impact (Residual)",
            "Residual Impact",
            "Post-Mitigation Impact",
            "Current Impact",
        ],
        "impact_type": [
            "Impact (Service unavailable, Endangerment, etc)",
            "Impact",
            "Severity",
        ],
        "mitigation": [
            "Proposed Solution",
            "Mitigation Strategy",
            "Mitigation",
            "Controls",
            "Response",
        ],
        "reason_postponed": ["Reason Postponed"],
        "date_postponed": ["Date Postponed", "Date Postponded"],
        "date_mitigated": ["Date Mitigated"],
        "mitigation_history": ["Mitigation History", "Notes"],
        "title": ["Risk", "Risk Title", "Title", "Issue", "Identified Risks"],
    }

    sheet_status = {
        "Review Queue": "In Review",
        "In-Progress": "In Progress",
        "Risks Exceptions": "Postponed",
        "Risks Mitigated": "Mitigated",
        "Risk Archive": "Archived",
    }

    def _as_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        return str(value).strip()

    def _find_header_row(rows: list[tuple]) -> int | None:
        for idx, row in enumerate(rows[:10]):
            row_text = " ".join(_as_text(cell) for cell in row)
            if "Description of Risk" in row_text or "Mitigation Strategy" in row_text:
                return idx
            if "Risk Owner" in row_text and "Priority" in row_text:
                return idx
        return None

    def _get_cell(row, header_map, key_aliases: list[str]) -> str:
        for key in key_aliases:
            normalized = _normalize_header(key)
            if normalized in header_map:
                value = row[header_map[normalized]]
                return _as_text(value)
        return ""

    inserted = 0
    now = datetime.utcnow().isoformat()
    with get_db_connection() as conn:
        for sheet_name in workbook.sheetnames:
            if sheet_name.strip().lower() == "kpi":
                continue

            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            meta_text = " ".join(_as_text(cell) for cell in rows[0])
            review_period = ""
            date_last_reviewed = ""
            next_review = ""
            for cell in rows[0]:
                text = _as_text(cell)
                if text.lower().startswith("review period"):
                    review_period = text.split(":", 1)[-1].strip() or review_period
                elif text.lower().startswith("date last reviewed"):
                    date_last_reviewed = text.split(":", 1)[-1].strip() or date_last_reviewed
                elif text.lower().startswith("next review"):
                    next_review = text.split(":", 1)[-1].strip() or next_review

            header_row_index = _find_header_row(rows)
            if header_row_index is None:
                continue

            headers = [
                _as_text(header) if header is not None else ""
                for header in rows[header_row_index]
            ]
            header_map = {_normalize_header(h): idx for idx, h in enumerate(headers)}

            status = sheet_status.get(sheet_name, "New")

            for row in rows[header_row_index + 1 :]:
                if not any(cell is not None and _as_text(cell) for cell in row):
                    continue

                title = _get_cell(row, header_map, header_aliases["title"])
                description = _get_cell(row, header_map, header_aliases["description"])
                owner = _get_cell(row, header_map, header_aliases["owner"])
                mitigation = _get_cell(row, header_map, header_aliases["mitigation"])
                likelihood = _get_cell(row, header_map, header_aliases["likelihood"])
                impact = _get_cell(row, header_map, header_aliases["impact"])
                likelihood_initial = _get_cell(
                    row, header_map, header_aliases["likelihood_initial"]
                )
                impact_initial = _get_cell(row, header_map, header_aliases["impact_initial"])
                likelihood_residual = _get_cell(
                    row, header_map, header_aliases["likelihood_residual"]
                )
                impact_residual = _get_cell(
                    row, header_map, header_aliases["impact_residual"]
                )
                impact_type = _get_cell(row, header_map, header_aliases["impact_type"])
                date_identified = _get_cell(row, header_map, header_aliases["date_identified"])
                priority_text = _get_cell(row, header_map, header_aliases["priority"])
                date_postponed = _get_cell(row, header_map, header_aliases["date_postponed"])
                reason_postponed = _get_cell(row, header_map, header_aliases["reason_postponed"])
                date_mitigated = _get_cell(row, header_map, header_aliases["date_mitigated"])
                mitigation_history = _get_cell(
                    row, header_map, header_aliases["mitigation_history"]
                )

                if not any(
                    [
                        title,
                        description,
                        mitigation,
                        impact_type,
                        owner,
                        likelihood,
                        impact,
                        likelihood_initial,
                        impact_initial,
                        likelihood_residual,
                        impact_residual,
                        date_identified,
                        priority_text,
                    ]
                ):
                    continue

                if not description and title:
                    description = title
                if not title and description:
                    title = description
                if not description or not title:
                    continue

                try:
                    priority = int(float(priority_text)) if priority_text else None
                except ValueError:
                    priority = None

                if not likelihood_residual and likelihood:
                    likelihood_residual = likelihood
                if not impact_residual and impact:
                    impact_residual = impact
                if not likelihood_initial and likelihood_residual:
                    likelihood_initial = likelihood_residual
                if not impact_initial and impact_residual:
                    impact_initial = impact_residual

                likelihood_value = likelihood_residual or likelihood_initial or ""
                impact_value = impact_residual or impact_initial or ""

                progress = 0
                if status in {"Mitigated", "Archived"}:
                    progress = 100
                elif status == "In Progress":
                    progress = 50

                conn.execute(
                    """
                    INSERT INTO risks (
                        title, description, likelihood, impact,
                        likelihood_initial, impact_initial,
                        likelihood_residual, impact_residual,
                        owner,
                        status, mitigation, progress, admin_notes,
                        date_identified, priority, impact_type, review_period,
                        date_last_reviewed, next_review, date_postponed,
                        reason_postponed, date_mitigated, mitigation_history,
                        source_sheet, created_at, updated_at
                    )
                    VALUES (
                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        ?, ?, ?, ?, ?, ?
                    )
                    """,
                    (
                        title,
                        description,
                        likelihood_value,
                        impact_value,
                        likelihood_initial,
                        impact_initial,
                        likelihood_residual,
                        impact_residual,
                        owner,
                        status,
                        mitigation,
                        progress,
                        "",
                        date_identified,
                        priority,
                        impact_type,
                        review_period,
                        date_last_reviewed,
                        next_review,
                        date_postponed,
                        reason_postponed,
                        date_mitigated,
                        mitigation_history,
                        sheet_name,
                        now,
                        now,
                    ),
                )
                inserted += 1
        conn.commit()

    _import_kpis(workbook)

    return inserted


def _rating_value(value: str | None) -> int | None:
    text = (value or "").strip().lower()
    if not text:
        return None
    mapping = {
        "very low": 1,
        "low": 2,
        "medium": 3,
        "high": 4,
        "very high": 5,
        "critical": 5,
    }
    if text in mapping:
        return mapping[text]
    # Accept numeric strings.
    try:
        n = int(float(text))
        if 0 <= n <= 10:
            return n
    except Exception:
        pass
    return None


def _rating_1_to_5(value: str | None) -> int | None:
    """Return a normalized 1..5 rating or None.

    Some legacy paths accept numeric strings; clamp anything outside the 1..5
    scale so admin analytics (heatmap/trends) never crash.
    """

    n = _rating_value(value)
    if n is None:
        return None
    if n <= 0:
        return None

    # Preferred scale is 1..5.
    if 1 <= n <= 5:
        return n

    # Legacy numeric imports sometimes used 1..10.
    # Map to 1..5 buckets: 1-2→1, 3-4→2, 5-6→3, 7-8→4, 9-10→5.
    if 1 <= n <= 10:
        return int((n + 1) // 2)

    return None


def _score_1_to_25(likelihood: int | None, impact: int | None) -> int | None:
    if likelihood is None or impact is None:
        return None
    try:
        l = int(likelihood)
        i = int(impact)
    except Exception:
        return None
    if l <= 0 or i <= 0:
        return None
    score = l * i
    if score < 1:
        return 1
    if score > 25:
        return 25
    return score


def _score_band(score: int | None) -> str:
    if score is None:
        return "Unscored"
    try:
        s = int(score)
    except Exception:
        return "Unscored"
    if s <= 0:
        return "Unscored"
    if s <= 3:
        return "Low"
    if s <= 8:
        return "Medium"
    if s <= 12:
        return "High"
    return "Critical"


def _parse_date(value: str | None) -> datetime | None:
    """Parse common date formats used in this app (ISO-ish strings)."""

    text = (value or "").strip()
    if not text:
        return None

    # Handle YYYY-MM-DD quickly.
    try:
        if len(text) == 10 and text[4] == "-" and text[7] == "-":
            return datetime.strptime(text, "%Y-%m-%d")
    except Exception:
        pass

    # Try ISO formats (with or without Z).
    try:
        cleaned = text.replace("Z", "+00:00")
        return datetime.fromisoformat(cleaned)
    except Exception:
        return None


def _today_ymd() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d")


def _compute_kpi_metrics(*, risks: list[sqlite3.Row]) -> dict[str, dict[str, str]]:
    completed_statuses = {"Mitigated", "Archived", "Closed"}
    total_risks = len(risks)
    completed_count = sum(1 for r in risks if (r["status"] or "").strip() in completed_statuses)

    metrics: dict[str, dict[str, str]] = {
        "total risks": {
            "value": str(total_risks),
            "notes": "Total risks currently in the system.",
        }
    }

    # Risk exposure: average residual score (likelihood×impact) where ratings exist.
    scored = 0
    residual_sum = 0
    for risk in risks:
        l = _rating_value(risk["likelihood_residual"])
        i = _rating_value(risk["impact_residual"])
        if l is None or i is None:
            continue
        scored += 1
        residual_sum += max(0, l) * max(0, i)
    if scored:
        avg = residual_sum / scored
        metrics["risk exposure"] = {
            "value": f"{avg:.1f}",
            "notes": f"Average residual risk score across {scored} risks with ratings.",
        }
    else:
        metrics["risk exposure"] = {
            "value": "—",
            "notes": "No residual likelihood/impact ratings available.",
        }

    # Mitigation effectiveness (prefer ratings-based score reduction; fallback to completion rate)
    scored_count = 0
    reduction_sum = 0.0
    score_delta_sum = 0.0
    improved_count = 0
    for risk in risks:
        initial_likelihood = _rating_value(risk["likelihood_initial"])
        initial_impact = _rating_value(risk["impact_initial"])
        residual_likelihood = _rating_value(risk["likelihood_residual"])
        residual_impact = _rating_value(risk["impact_residual"])
        if None in (initial_likelihood, initial_impact, residual_likelihood, residual_impact):
            continue

        initial_score = max(0, initial_likelihood) * max(0, initial_impact)
        residual_score = max(0, residual_likelihood) * max(0, residual_impact)
        if initial_score <= 0:
            continue

        scored_count += 1
        reduction = (initial_score - residual_score) / initial_score
        reduction_sum += reduction
        score_delta_sum += (residual_score - initial_score)
        if residual_score < initial_score:
            improved_count += 1

    if scored_count:
        avg_reduction = reduction_sum / scored_count
        avg_delta = score_delta_sum / scored_count
        metrics["risk mitigation effectiveness"] = {
            "value": f"{avg_reduction * 100:.0f}%",
            "notes": (
                f"Average reduction in risk score (likelihood×impact) across {scored_count} risks; "
                f"{improved_count} improved. Avg score change: {avg_delta:.1f}."
            ),
        }
    else:
        completion_rate = (completed_count / total_risks * 100.0) if total_risks else 0.0
        metrics["risk mitigation effectiveness"] = {
            "value": f"{completion_rate:.0f}%",
            "notes": (
                "Fallback metric: percent of risks in a completed status (Mitigated/Archived/Closed). "
                "Add initial/residual likelihood & impact ratings to enable score-based effectiveness."
            ),
        }

    # Response timeliness (prefer time-to-mitigate; fallback to average age of open risks)
    total_days = 0
    count_days = 0
    for risk in risks:
        if (risk["status"] or "").strip() not in completed_statuses:
            continue
        identified = _parse_date(risk["date_identified"])
        mitigated = _parse_date(risk["date_mitigated"])
        if not identified or not mitigated:
            continue
        total_days += (mitigated - identified).days
        count_days += 1

    if count_days:
        average_days = total_days / count_days
        metrics["risk response timeliness"] = {
            "value": f"{average_days:.1f} days",
            "notes": f"Average time from identification to mitigation across {count_days} completed risks.",
        }
    else:
        today = datetime.utcnow()
        open_days_total = 0
        open_days_count = 0
        for risk in risks:
            if (risk["status"] or "").strip() in completed_statuses:
                continue
            identified = _parse_date(risk["date_identified"])
            if not identified:
                continue
            open_days_total += (today - identified).days
            open_days_count += 1
        if open_days_count:
            avg_age = open_days_total / open_days_count
            metrics["risk response timeliness"] = {
                "value": f"{avg_age:.1f} days",
                "notes": f"Fallback metric: average age of open risks across {open_days_count} risks.",
            }
        else:
            metrics["risk response timeliness"] = {
                "value": "—",
                "notes": "No usable dates found (date_identified/date_mitigated).",
            }

    return metrics


def _import_kpis(workbook) -> None:
    """Import KPI sheet rows into the kpis table (best-effort)."""

    try:
        sheet_name = next((n for n in workbook.sheetnames if str(n).strip().lower() == "kpi"), "")
        if not sheet_name:
            return
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return

        now = datetime.utcnow().isoformat()
        to_insert: list[tuple[str, str, str]] = []
        for row in rows:
            if not row:
                continue
            name = str(row[0] or "").strip()
            if not name:
                continue
            value = str(row[1] or "").strip() if len(row) > 1 else ""
            notes = str(row[2] or "").strip() if len(row) > 2 else ""
            to_insert.append((name, value, notes))

        if not to_insert:
            return

        with get_db_connection() as conn:
            for name, value, notes in to_insert:
                conn.execute(
                    """
                    INSERT INTO kpis (name, value, notes, source_sheet, created_at)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (name, value, notes, sheet_name, now),
                )
            conn.commit()
    except Exception:
        return


def _median(values: list[float]) -> float | None:
    values = sorted(v for v in values if v is not None)
    if not values:
        return None
    mid = len(values) // 2
    if len(values) % 2:
        return float(values[mid])
    return (float(values[mid - 1]) + float(values[mid])) / 2.0


def _build_admin_risks_query(
    *,
    status_filter: str,
    severity_filter: str,
    assignee_filter: str,
    query_text: str,
    current_user: str,
) -> tuple[str, list[str]]:
    query = "SELECT * FROM risks"
    filters: list[str] = []
    params: list[str] = []

    # Hide soft-deleted risks in normal admin lists.
    filters.append(_not_deleted_condition())

    if status_filter != "all":
        filters.append("LOWER(TRIM(status)) = ?")
        params.append(status_filter.lower())
    if severity_filter != "all":
        filters.append("LOWER(TRIM(COALESCE(severity, ''))) = ?")
        params.append(severity_filter.lower())

    if assignee_filter == "me":
        cond_sql, cond_params = _assignee_me_condition(current_user)
        if cond_sql:
            filters.append(cond_sql)
            params.extend(cond_params)
        else:
            # If we don't know who "me" is, return no rows.
            filters.append("1 = 0")
    elif assignee_filter == "unassigned":
        filters.append("TRIM(COALESCE(assigned_to, '')) = ''")
    elif assignee_filter != "all":
        filters.append("LOWER(TRIM(COALESCE(assigned_to, ''))) = ?")
        params.append(assignee_filter.lower())

    if query_text:
        filters.append("(LOWER(title) LIKE ? OR LOWER(description) LIKE ?)")
        like = f"%{query_text.lower()}%"
        params.extend([like, like])

    if filters:
        query += " WHERE " + " AND ".join(filters)
    query += " ORDER BY updated_at DESC"
    return query, params


def _build_admin_risks_where(
    *,
    status_filter: str,
    severity_filter: str,
    assignee_filter: str,
    query_text: str,
    current_user: str,
) -> tuple[str, list[str]]:
    filters: list[str] = []
    params: list[str] = []

    # Hide soft-deleted risks in normal admin lists.
    filters.append(_not_deleted_condition())

    if status_filter != "all":
        filters.append("LOWER(TRIM(status)) = ?")
        params.append(status_filter.lower())
    if severity_filter != "all":
        filters.append("LOWER(TRIM(COALESCE(severity, ''))) = ?")
        params.append(severity_filter.lower())

    if assignee_filter == "me":
        cond_sql, cond_params = _assignee_me_condition(current_user)
        if cond_sql:
            filters.append(cond_sql)
            params.extend(cond_params)
        else:
            filters.append("1 = 0")
    elif assignee_filter == "unassigned":
        filters.append("TRIM(COALESCE(assigned_to, '')) = ''")
    elif assignee_filter != "all":
        filters.append("LOWER(TRIM(COALESCE(assigned_to, ''))) = ?")
        params.append(assignee_filter.lower())

    if query_text:
        filters.append("(LOWER(title) LIKE ? OR LOWER(description) LIKE ?)")
        like = f"%{query_text.lower()}%"
        params.extend([like, like])

    if filters:
        return " WHERE " + " AND ".join(filters), params
    return "", params


def _build_admin_export_where(
    *,
    status_filter: str,
    severity_filter: str,
    assignee_filter: str,
    owner_filter: str,
    query_text: str,
    current_user: str,
    date_field: str,
    date_from: str,
    date_to: str,
) -> tuple[str, list[str]]:
    filters: list[str] = []
    params: list[str] = []

    # Hide soft-deleted risks in normal exports.
    filters.append(_not_deleted_condition())

    # Reuse the same semantics as the admin dashboard filters.
    if status_filter != "all":
        filters.append("LOWER(TRIM(status)) = ?")
        params.append(status_filter.lower())
    if severity_filter != "all":
        filters.append("LOWER(TRIM(COALESCE(severity, ''))) = ?")
        params.append(severity_filter.lower())

    if assignee_filter == "me":
        cond_sql, cond_params = _assignee_me_condition(current_user)
        if cond_sql:
            filters.append(cond_sql)
            params.extend(cond_params)
        else:
            filters.append("1 = 0")
    elif assignee_filter == "unassigned":
        filters.append("TRIM(COALESCE(assigned_to, '')) = ''")
    elif assignee_filter != "all":
        filters.append("LOWER(TRIM(COALESCE(assigned_to, ''))) = ?")
        params.append(assignee_filter.lower())

    if owner_filter != "all":
        filters.append("LOWER(TRIM(COALESCE(owner, ''))) = ?")
        params.append(owner_filter.lower())

    if query_text:
        filters.append("(LOWER(title) LIKE ? OR LOWER(description) LIKE ?)")
        like = f"%{query_text.lower()}%"
        params.extend([like, like])

    allowed_date_fields = {
        "created_at": "created_at",
        "updated_at": "updated_at",
        "closed_at": "closed_at",
        "date_identified": "date_identified",
    }
    safe_date_col = allowed_date_fields.get((date_field or "").strip(), "created_at")

    # NOTE: Dates in this app are stored as text; DATE() works for ISO-like strings.
    if date_from:
        filters.append(f"DATE({safe_date_col}) >= DATE(?)")
        params.append(date_from)
    if date_to:
        filters.append(f"DATE({safe_date_col}) <= DATE(?)")
        params.append(date_to)

    if filters:
        return " WHERE " + " AND ".join(filters), params
    return "", params


def _match_kpi_metric(name: str) -> str | None:
    text = " ".join((name or "").strip().lower().split())
    if not text:
        return None
    if "exposure" in text:
        return "risk exposure"
    if "effectiveness" in text or ("mitigation" in text and "effect" in text):
        return "risk mitigation effectiveness"
    if "timeliness" in text or "turnaround" in text or ("response" in text and "time" in text):
        return "risk response timeliness"
    return None


def _compute_severity(
    likelihood: str | None,
    impact: str | None,
    fallback: str = "Medium",
) -> str:
    likelihood_value = _rating_value(likelihood)
    impact_value = _rating_value(impact)
    if likelihood_value is None or impact_value is None:
        return fallback

    score = max(0, likelihood_value) * max(0, impact_value)
    if score <= 3:
        return "Low"
    if score <= 8:
        return "Medium"
    if score <= 12:
        return "High"
    return "Critical"


def _log_event(
    conn: sqlite3.Connection,
    *,
    risk_id: int,
    event_type: str,
    field: str | None = None,
    old_value: str | None = None,
    new_value: str | None = None,
    actor: str | None = None,
) -> None:
    now = datetime.utcnow().isoformat()
    conn.execute(
        """
        INSERT INTO risk_events (risk_id, event_type, field, old_value, new_value, actor, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            risk_id,
            event_type,
            field,
            (old_value or "")[:500],
            (new_value or "")[:500],
            (actor or "")[:200],
            now,
        ),
    )


def _allowed_upload(filename: str) -> bool:
    allowed = {
        ".png",
        ".jpg",
        ".jpeg",
        ".gif",
        ".webp",
        ".pdf",
        ".txt",
        ".log",
        ".csv",
        ".xlsx",
        ".docx",
        ".msg",
        ".eml",
    }
    suffix = Path(filename).suffix.lower()
    return bool(suffix) and suffix in allowed


def _risk_upload_path(risk_id: int) -> Path:
    return UPLOAD_DIR / str(risk_id)


def _risk_archive_path(risk_id: int) -> Path:
    return ARCHIVE_DIR / str(risk_id)


def _split_kpis(kpis: list[sqlite3.Row]) -> tuple[list[dict], list[dict]]:
    headers = {
        "potential kpis",
        "chosen kpis",
        "how i'll mesure",
        "how i'll measure",
        "formula",
    }
    potential: list[dict] = []
    chosen: list[dict] = []
    for kpi in kpis:
        name = (kpi["name"] or "").strip()
        if not name:
            continue
        if name.lower() in headers:
            continue
        value = (kpi["value"] or "").strip()
        notes = (kpi["notes"] or "").strip()
        is_formula = notes.startswith("=") or ("IF(" in notes.upper())
        if value.startswith("="):
            is_formula = True
        notes = "" if is_formula else notes
        if len(notes) > 140:
            notes = notes[:137].rstrip() + "..."
        if value or notes:
            chosen.append(
                {"name": name, "value": value, "notes": notes, "is_formula": is_formula}
            )
        else:
            potential.append({"name": name})
    return potential, chosen


def _format_int(value: int) -> str:
    return f"{int(value):,}"


def _format_percent(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return "0%"
    return f"{(numerator / denominator) * 100.0:.0f}%"


def _compute_program_kpis(
    conn: sqlite3.Connection,
    *,
    where_sql: str,
    params: list[str],
) -> list[dict[str, str]]:
    """Compute always-valid KPIs directly from the DB.

    These are intended to be stable operational metrics for the Risk Ticketing app,
    not spreadsheet formulas.
    """

    completed_statuses = ["mitigated", "archived", "closed"]
    completed_placeholders = ",".join(["?"] * len(completed_statuses))
    open_condition = f"LOWER(TRIM(COALESCE(status, ''))) NOT IN ({completed_placeholders})"
    completed_condition = f"LOWER(TRIM(COALESCE(status, ''))) IN ({completed_placeholders})"

    total = int(conn.execute(f"SELECT COUNT(*) FROM risks{where_sql}", params).fetchone()[0])
    open_total = int(
        conn.execute(
            f"SELECT COUNT(*) FROM risks{_where_join(where_sql, open_condition)}",
            params + completed_statuses,
        ).fetchone()[0]
    )
    completed_total = int(
        conn.execute(
            f"SELECT COUNT(*) FROM risks{_where_join(where_sql, completed_condition)}",
            params + completed_statuses,
        ).fetchone()[0]
    )
    high_critical_open = int(
        conn.execute(
            f"""
            SELECT COUNT(*)
            FROM risks
            {_where_join(where_sql, f"LOWER(TRIM(COALESCE(severity,''))) IN ('high','critical') AND {open_condition}")}
            """.strip(),
            params + completed_statuses,
        ).fetchone()[0]
    )
    unassigned_open = int(
        conn.execute(
            f"""
            SELECT COUNT(*)
            FROM risks
            {_where_join(where_sql, f"TRIM(COALESCE(assigned_to, '')) = '' AND {open_condition}")}
            """.strip(),
            params + completed_statuses,
        ).fetchone()[0]
    )

    # Median first-touch hours (from risk_events). We treat the first 'update' event as first-touch.
    touch_rows = conn.execute(
        f"""
        SELECT r.id as id,
               r.status as status,
               r.created_at as created_at,
               MIN(e.created_at) as first_update_at
        FROM risks r
        LEFT JOIN risk_events e
            ON e.risk_id = r.id
           AND e.event_type = 'update'
        {where_sql}
        GROUP BY r.id
        """.strip(),
        params,
    ).fetchall()

    first_touch_hours: list[float] = []
    first_touch_breaches_24h = 0
    open_seen = 0
    for row in touch_rows:
        created = _parse_date(row["created_at"])
        touched = _parse_date(row["first_update_at"])
        status_lower = str(row["status"] or "").strip().lower()
        is_open = status_lower not in set(completed_statuses)
        if is_open:
            open_seen += 1
        if created and touched and touched >= created:
            delta_hours = (touched - created).total_seconds() / 3600.0
            first_touch_hours.append(delta_hours)
            if is_open and delta_hours > 24.0:
                first_touch_breaches_24h += 1

    median_first_touch = _median(first_touch_hours)

    # Assignment SLAs (from risk_events field='assigned_to').
    assign_rows = conn.execute(
        f"""
        SELECT r.id as id,
               r.created_at as created_at,
               MIN(e.created_at) as assigned_at
        FROM risks r
        LEFT JOIN risk_events e
            ON e.risk_id = r.id
           AND e.event_type = 'update'
           AND e.field = 'assigned_to'
           AND TRIM(COALESCE(e.new_value, '')) <> ''
        {where_sql}
        GROUP BY r.id
        """.strip(),
        params,
    ).fetchall()

    assign_hours: list[float] = []
    assigned_within_24h = 0
    assigned_count = 0
    for row in assign_rows:
        created = _parse_date(row["created_at"])
        assigned_at = _parse_date(row["assigned_at"])
        if created and assigned_at and assigned_at >= created:
            assigned_count += 1
            delta_hours = (assigned_at - created).total_seconds() / 3600.0
            assign_hours.append(delta_hours)
            if delta_hours <= 24.0:
                assigned_within_24h += 1

    median_assign = _median(assign_hours)

    # Aging + review hygiene (computed for open risks).
    open_over_30d = int(
        conn.execute(
            f"""
            SELECT COUNT(*)
            FROM risks
            {_where_join(where_sql, f"{open_condition} AND (julianday('now') - julianday(replace(created_at, 'T', ' '))) > 30")}
            """.strip(),
            params + completed_statuses,
        ).fetchone()[0]
    )

    open_review_rows = conn.execute(
        f"""
        SELECT created_at, next_review, assigned_to
        FROM risks
        {_where_join(where_sql, open_condition)}
        """.strip(),
        params + completed_statuses,
    ).fetchall()

    now = datetime.utcnow()
    overdue_reviews = 0
    unassigned_breaches_24h = 0
    for row in open_review_rows:
        created = _parse_date(row["created_at"])
        if created and (now - created).total_seconds() > 24.0 * 3600.0:
            if not str(row["assigned_to"] or "").strip():
                unassigned_breaches_24h += 1

        next_review = _parse_date(row["next_review"])
        if next_review and next_review.date() < now.date():
            overdue_reviews += 1

    # Median time-to-done days (Closed preferred, else Date Mitigated).
    done_rows = conn.execute(
        f"""
        SELECT created_at,
               COALESCE(NULLIF(TRIM(closed_at), ''), NULLIF(TRIM(date_mitigated), '')) as done_at
        FROM risks
        {_where_join(where_sql, completed_condition)}
        """.strip(),
        params + completed_statuses,
    ).fetchall()

    done_days: list[float] = []
    for row in done_rows:
        created = _parse_date(row["created_at"])
        done_at = _parse_date(row["done_at"])
        if created and done_at and done_at >= created:
            done_days.append((done_at - created).total_seconds() / (3600.0 * 24.0))
    median_done_days = _median(done_days)

    cards: list[dict[str, str]] = [
        {
            "name": "Total risks",
            "value": _format_int(total),
            "notes": "All risks in the current view.",
        },
        {
            "name": "Open risks",
            "value": _format_int(open_total),
            "notes": "Not Mitigated/Archived/Closed.",
        },
        {
            "name": "Open > 30 days",
            "value": _format_int(open_over_30d),
            "notes": "Open risks older than 30 days (based on created_at).",
        },
        {
            "name": "Completed rate",
            "value": _format_percent(completed_total, max(1, total)),
            "notes": "Percent Mitigated/Archived/Closed.",
        },
        {
            "name": "High/Critical open",
            "value": _format_int(high_critical_open),
            "notes": "High/critical severity that are still open.",
        },
        {
            "name": "Unassigned open",
            "value": _format_int(unassigned_open),
            "notes": "Open risks missing an assignee.",
        },
        {
            "name": "Overdue reviews",
            "value": _format_int(overdue_reviews),
            "notes": "Open risks where Next Review is before today.",
        },
        {
            "name": "Median first touch",
            "value": (f"{median_first_touch:.1f}h" if median_first_touch is not None else "—"),
            "notes": "Median time from create → first update event.",
        },
        {
            "name": "First-touch breaches",
            "value": _format_int(first_touch_breaches_24h),
            "notes": "Open risks with first touch > 24h.",
        },
        {
            "name": "Median time to assign",
            "value": (f"{median_assign:.1f}h" if median_assign is not None else "—"),
            "notes": "Median time from create → first assignee set.",
        },
        {
            "name": "Assigned within 24h",
            "value": _format_percent(assigned_within_24h, max(1, assigned_count)),
            "notes": "Percent of assigned risks whose first assignment happened within 24h.",
        },
        {
            "name": "Unassigned breaches",
            "value": _format_int(unassigned_breaches_24h),
            "notes": "Open risks unassigned for > 24h.",
        },
        {
            "name": "Median time to done",
            "value": (f"{median_done_days:.1f}d" if median_done_days is not None else "—"),
            "notes": "Median create → closed/mitigated.",
        },
    ]

    return cards


@app.before_request
def setup() -> None:
    if not app.config["DB_INITIALIZED"]:
        init_db()
        seed_from_excel()
        app.config["DB_INITIALIZED"] = True
    if "admin_import" not in app.view_functions:
        def _admin_import_stub():
            flash("Import is not available. Please restart the service.", "warning")
            return redirect(url_for("admin_dashboard"))

        app.add_url_rule(
            "/admin/import",
            endpoint="admin_import",
            view_func=_admin_import_stub,
            methods=["POST"],
        )


@app.route("/")
def index():
    status_filter = request.args.get("status", "all").strip()
    owner_filter = request.args.get("owner", "all").strip()
    sheet_filter = request.args.get("sheet", "all").strip()
    page, per_page = _parse_pagination(default_per_page=50)
    query = "SELECT * FROM risks"
    filters = []
    params = []

    # Exclude soft-deleted risks from the public list.
    filters.append(_not_deleted_condition())

    if status_filter != "all":
        filters.append("LOWER(TRIM(status)) = ?")
        params.append(status_filter.lower())
    if owner_filter != "all":
        filters.append("LOWER(TRIM(owner)) = ?")
        params.append(owner_filter.lower())
    if sheet_filter != "all":
        filters.append("LOWER(TRIM(source_sheet)) = ?")
        params.append(sheet_filter.lower())
    if filters:
        query += " WHERE " + " AND ".join(filters)
    where_sql = (" WHERE " + " AND ".join(filters)) if filters else ""
    query += " ORDER BY updated_at DESC"

    paged_query = query + " LIMIT ? OFFSET ?"
    offset = (page - 1) * per_page

    with get_db_connection() as conn:
        total = int(conn.execute(f"SELECT COUNT(*) FROM risks{where_sql}", params).fetchone()[0])
        pager = _pager(total=total, page=page, per_page=per_page)
        offset = (int(pager["page"]) - 1) * per_page
        risks = conn.execute(paged_query, params + [per_page, offset]).fetchall()
        statuses = [
            row["status"]
            for row in conn.execute(
                f"SELECT DISTINCT TRIM(status) as status FROM risks WHERE {_not_deleted_condition()} AND status IS NOT NULL AND TRIM(status) <> ''"
            )
        ]
        owners = [
            row["owner"]
            for row in conn.execute(
                f"SELECT DISTINCT TRIM(owner) as owner FROM risks WHERE {_not_deleted_condition()} AND owner IS NOT NULL AND TRIM(owner) <> ''"
            )
        ]
        sheets = [
            row["source_sheet"]
            for row in conn.execute(
                f"SELECT DISTINCT TRIM(source_sheet) as source_sheet FROM risks WHERE {_not_deleted_condition()} AND source_sheet IS NOT NULL AND TRIM(source_sheet) <> ''"
            )
        ]
        program_kpis = _compute_program_kpis(conn, where_sql=where_sql, params=params)
        summary = conn.execute(
            """
            SELECT status, COUNT(*) as count
            FROM risks
            WHERE TRIM(COALESCE(deleted_at, '')) = ''
            GROUP BY status
            """
        ).fetchall()

    query_args = {
        "status": status_filter,
        "owner": owner_filter,
        "sheet": sheet_filter,
        "per_page": int(pager["per_page"]),
    }
    if pager.get("has_prev"):
        pager["prev_url"] = url_for("index", **query_args, page=int(pager["page"]) - 1)
    if pager.get("has_next"):
        pager["next_url"] = url_for("index", **query_args, page=int(pager["page"]) + 1)

    return render_template(
        "index.html",
        risks=risks,
        statuses=sorted(set(statuses)),
        owners=sorted(set(owners)),
        sheets=sorted(set(sheets)),
        kpi_cards=program_kpis,
        summary=summary,
        status_filter=status_filter,
        owner_filter=owner_filter,
        sheet_filter=sheet_filter,
        pagination=pager,
        query_args=query_args,
    )


@app.route("/_health")
def health():
    return {
        "version": app.config.get("APP_VERSION"),
        "admin_import": "admin_import" in app.view_functions,
    }


@app.route("/submit", methods=["GET", "POST"])
def submit():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        severity = request.form.get("severity", "").strip()
        likelihood_initial = request.form.get("likelihood_initial", "").strip()
        impact_initial = request.form.get("impact_initial", "").strip()
        likelihood_residual = request.form.get("likelihood_residual", "").strip()
        impact_residual = request.form.get("impact_residual", "").strip()
        likelihood = request.form.get("likelihood", "").strip()
        impact = request.form.get("impact", "").strip()
        owner = request.form.get("owner", "").strip()
        priority_text = request.form.get("priority", "").strip()
        impact_type = request.form.get("impact_type", "").strip()

        if not title or not description:
            flash("Title and description are required.", "danger")
            return render_template("submit.html")

        now = datetime.utcnow().isoformat()
        try:
            priority = int(float(priority_text)) if priority_text else None
        except ValueError:
            priority = None

        if not likelihood_residual and likelihood:
            likelihood_residual = likelihood
        if not impact_residual and impact:
            impact_residual = impact
        if not likelihood_initial and likelihood_residual:
            likelihood_initial = likelihood_residual
        if not impact_initial and impact_residual:
            impact_initial = impact_residual

        likelihood_value = likelihood_residual or likelihood_initial
        impact_value = impact_residual or impact_initial
        if not severity:
            severity = _compute_severity(likelihood_value, impact_value)
        if severity not in SEVERITY_OPTIONS:
            severity = _compute_severity(likelihood_value, impact_value)

        if owner:
            owner_email = _normalize_email(owner)
            if not owner_email:
                flash("Suggested Owner must be an email address (e.g. name@hdh.org).", "warning")
                return render_template("submit.html")
            owner = owner_email

        actor = _current_actor()
        with get_db_connection() as conn:
            cursor = conn.execute(
                """
                INSERT INTO risks (
                    title, description, likelihood, impact,
                    likelihood_initial, impact_initial,
                    likelihood_residual, impact_residual,
                    owner,
                    status, mitigation, progress, admin_notes,
                    date_identified, priority, impact_type, review_period,
                    severity, assigned_to, created_by, updated_by,
                    created_at, updated_at
                )
                VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?,
                    ?, 'New', '', 0, '',
                    ?, ?, ?, '',
                    ?, ?, ?, ?,
                    ?, ?
                )
                """,
                (
                    title,
                    description,
                    likelihood_value,
                    impact_value,
                    likelihood_initial,
                    impact_initial,
                    likelihood_residual,
                    impact_residual,
                    owner,
                    now,
                    priority,
                    impact_type,
                    severity,
                    "",
                    actor,
                    actor,
                    now,
                    now,
                ),
            )
            risk_id = int(cursor.lastrowid)
            _log_event(
                conn,
                risk_id=risk_id,
                event_type="create",
                actor=actor,
            )
            conn.commit()

            # Notifications: new critical risk
            try:
                notify_to = _parse_recipients(os.getenv("NOTIFY_CRITICAL_TO", ""))
                if (severity or "").strip() == "Critical" and notify_to:
                    base_url = os.getenv("APP_BASE_URL", "").strip() or request.url_root.rstrip("/")
                    link = f"{base_url}{url_for('admin_risk_detail', risk_id=risk_id)}"
                    sent = _send_email(
                        subject=f"[CRITICAL RISK] {title}",
                        html_body=(
                            f"<p>A new <strong>Critical</strong> risk was submitted.</p>"
                            f"<p><strong>Title:</strong> {title}</p>"
                            f"<p><strong>Description:</strong><br/>{(description or '').replace('\n','<br/>')}</p>"
                            f"<p><a href=\"{link}\">Open in Admin Dashboard</a></p>"
                        ),
                        to_addrs=notify_to,
                    )
                    if sent:
                        _log_event(
                            conn,
                            risk_id=risk_id,
                            event_type="notify",
                            field="email",
                            old_value="",
                            new_value=",".join(notify_to),
                            actor=actor,
                        )
                        conn.commit()
            except Exception:
                pass

            # Notifications: suggested owner (if provided)
            try:
                owner_email = _normalize_email(owner)
                if owner_email and _env_bool("NOTIFY_OWNER_ON_SUBMIT_ENABLED", True):
                    base_url = os.getenv("APP_BASE_URL", "").strip() or request.url_root.rstrip("/")
                    link = f"{base_url}{url_for('risk_detail', risk_id=risk_id)}"
                    sent = _send_email(
                        subject=f"Risk Submitted: {title}",
                        html_body=(
                            f"<p>A risk was submitted and you were listed as the suggested owner.</p>"
                            f"<p><strong>Title:</strong> {title}</p>"
                            f"<p><strong>Severity:</strong> {severity or 'Unspecified'}</p>"
                            f"<p><a href=\"{link}\">Open risk</a></p>"
                        ),
                        to_addrs=[owner_email],
                    )
                    if sent:
                        _log_event(
                            conn,
                            risk_id=risk_id,
                            event_type="notify",
                            field="owner_submit",
                            old_value="",
                            new_value=owner_email,
                            actor=actor,
                        )
                        conn.commit()
            except Exception:
                pass

        flash("Risk submitted successfully.", "success")
        return redirect(url_for("index"))

    return render_template("submit.html")


@app.route("/risk/<int:risk_id>")
def risk_detail(risk_id: int):
    try:
        events_page = int(float(request.args.get("events_page", "1")))
    except ValueError:
        events_page = 1
    events_page = max(1, events_page)
    events_per_page = 25

    with get_db_connection() as conn:
        risk = conn.execute(
            f"SELECT * FROM risks WHERE id = ? AND {_not_deleted_condition()}",
            (risk_id,),
        ).fetchone()

        events_total = int(
            conn.execute(
                "SELECT COUNT(*) FROM risk_events WHERE risk_id = ?",
                (risk_id,),
            ).fetchone()[0]
        )
        events_pagination = _pager(total=events_total, page=events_page, per_page=events_per_page)
        events_offset = (int(events_pagination["page"]) - 1) * events_per_page
        events = conn.execute(
            "SELECT * FROM risk_events WHERE risk_id = ? ORDER BY id DESC LIMIT ? OFFSET ?",
            (risk_id, events_per_page, events_offset),
        ).fetchall()

        attachments = conn.execute(
            "SELECT * FROM risk_attachments WHERE risk_id = ? ORDER BY id DESC",
            (risk_id,),
        ).fetchall()
    if risk is None:
        flash("Risk not found.", "warning")
        return redirect(url_for("index"))

    events_query_args = request.args.to_dict(flat=True)
    if events_pagination["has_prev"]:
        prev_args = dict(events_query_args)
        prev_args["events_page"] = int(events_pagination["page"]) - 1
        events_pagination["prev_url"] = url_for("risk_detail", risk_id=risk_id, **prev_args)
    if events_pagination["has_next"]:
        next_args = dict(events_query_args)
        next_args["events_page"] = int(events_pagination["page"]) + 1
        events_pagination["next_url"] = url_for("risk_detail", risk_id=risk_id, **next_args)

    return render_template(
        "risk_detail.html",
        risk=risk,
        attachments=attachments,
        events=events,
        events_pagination=events_pagination,
    )


@app.route("/risk/<int:risk_id>/attachment/<int:attachment_id>")
@login_required
def risk_attachment_download(risk_id: int, attachment_id: int):
    with get_db_connection() as conn:
        attachment = conn.execute(
            "SELECT * FROM risk_attachments WHERE id = ? AND risk_id = ?",
            (attachment_id, risk_id),
        ).fetchone()
    if attachment is None:
        abort(404)

    storage_name = attachment["storage_name"]
    original_name = attachment["original_name"] or storage_name
    is_archived = int(attachment["is_archived"] or 0) == 1
    file_path = (_risk_archive_path(risk_id) if is_archived else _risk_upload_path(risk_id)) / storage_name
    if not file_path.exists():
        abort(404)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=original_name,
    )


@app.route("/admin/attachments/archive_closed", methods=["POST"])
def admin_archive_closed_attachments():
    if not require_admin():
        return redirect(url_for("admin"))

    try:
        days = int(float((request.form.get("days") or "").strip() or str(ARCHIVE_AFTER_DAYS)))
    except ValueError:
        days = ARCHIVE_AFTER_DAYS

    now = datetime.utcnow()
    cutoff = now.timestamp() - max(0, days) * 86400
    cutoff_iso = datetime.utcfromtimestamp(cutoff).isoformat()

    moved = 0
    flagged = 0
    actor = _current_actor()
    now_iso = now.isoformat()

    with get_db_connection() as conn:
        # Only archive attachments for risks that are Closed and have a closed_at older than cutoff
        rows = conn.execute(
            """
            SELECT a.*
            FROM risk_attachments a
            JOIN risks r ON r.id = a.risk_id
            WHERE (a.is_archived IS NULL OR a.is_archived = 0)
              AND (r.status = 'Closed')
              AND (r.closed_at IS NOT NULL AND r.closed_at != '')
              AND r.closed_at < ?
            ORDER BY a.risk_id, a.id
            """,
            (cutoff_iso,),
        ).fetchall()

        for a in rows:
            rid = int(a["risk_id"])
            storage_name = a["storage_name"]
            original_name = a["original_name"] or storage_name

            src = _risk_upload_path(rid) / storage_name
            dst_dir = _risk_archive_path(rid)
            dst_dir.mkdir(parents=True, exist_ok=True)
            dst = dst_dir / storage_name

            try:
                if src.exists() and src.is_file():
                    src.replace(dst)
                    moved += 1
            except Exception:
                # If move fails, still mark archived to reduce hot storage (ops can reconcile)
                pass

            conn.execute(
                "UPDATE risk_attachments SET is_archived = 1, archived_at = ? WHERE id = ?",
                (now_iso, int(a["id"])),
            )
            _log_event(
                conn,
                risk_id=rid,
                event_type="attachment_archive",
                field="attachment",
                old_value=original_name,
                new_value="archived",
                actor=actor,
            )
            flagged += 1

        conn.commit()

    flash(f"Archived {flagged} attachment record(s) ({moved} file(s) moved).", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin", methods=["GET", "POST"])
def admin():
    # If admin password mode is disabled, /admin becomes an access-check page.
    if not app.config.get("ADMIN_PASSWORD_ENABLED", True):
        if is_admin():
            return redirect(url_for("admin_dashboard"))
        return render_template("admin_login.html", password_enabled=False), 403

    if request.method == "POST":
        password = request.form.get("password", "")
        if password == app.config["ADMIN_PASSWORD"]:
            session["is_admin"] = True
            who = _normalize_email(request.form.get("who") or "")
            if who:
                session["admin_user_email"] = who
            else:
                session.pop("admin_user_email", None)
            flash("Welcome, admin.", "success")
            return redirect(url_for("admin_dashboard"))
        flash("Invalid admin password.", "danger")

    if is_admin():
        return redirect(url_for("admin_dashboard"))

    return render_template("admin_login.html", password_enabled=True)


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    session.pop("admin_user_email", None)
    flash("Logged out.", "info")
    return redirect(url_for("index"))


@app.route("/admin/dashboard")
def admin_dashboard():
    if not require_admin():
        return redirect(url_for("admin"))

    status_filter = request.args.get("status", "all").strip()
    severity_filter = request.args.get("severity", "all").strip()
    assignee_filter = request.args.get("assigned", "all").strip()
    query_text = request.args.get("q", "").strip()
    page, per_page = _parse_pagination(default_per_page=50)

    current_user = _current_user_email()
    where_sql, params = _build_admin_risks_where(
        status_filter=status_filter,
        severity_filter=severity_filter,
        assignee_filter=assignee_filter,
        query_text=query_text,
        current_user=current_user,
    )

    list_query = f"SELECT * FROM risks{where_sql} ORDER BY updated_at DESC LIMIT ? OFFSET ?"

    with get_db_connection() as conn:
        total = int(conn.execute(f"SELECT COUNT(*) FROM risks{where_sql}", params).fetchone()[0])
        pager = _pager(total=total, page=page, per_page=per_page)
        offset = (int(pager["page"]) - 1) * per_page
        risks = conn.execute(list_query, params + [per_page, offset]).fetchall()
        statuses = [
            row["status"]
            for row in conn.execute(
                f"SELECT DISTINCT TRIM(status) as status FROM risks WHERE {_not_deleted_condition()} AND status IS NOT NULL AND TRIM(status) <> ''"
            )
        ]
        severities = [
            row["severity"]
            for row in conn.execute(
                f"SELECT DISTINCT TRIM(severity) as severity FROM risks WHERE {_not_deleted_condition()} AND severity IS NOT NULL AND TRIM(severity) <> ''"
            )
        ]
        assignees = [
            row["assigned_to"]
            for row in conn.execute(
                f"SELECT DISTINCT TRIM(assigned_to) as assigned_to FROM risks WHERE {_not_deleted_condition()} AND assigned_to IS NOT NULL AND TRIM(assigned_to) <> ''"
            )
        ]
        allowed_assignees, _assignee_source = _allowed_assignee_emails()
        if allowed_assignees:
            assignees = allowed_assignees
        program_kpis = _compute_program_kpis(conn, where_sql=where_sql, params=params)
        summary = conn.execute(
            """
            SELECT status, COUNT(*) as count
            FROM risks
            WHERE TRIM(COALESCE(deleted_at, '')) = ''
            GROUP BY status
            """
        ).fetchall()

        severity_summary = conn.execute(
            """
            SELECT COALESCE(severity, 'Unspecified') as severity, COUNT(*) as count
            FROM risks
            WHERE TRIM(COALESCE(deleted_at, '')) = ''
            GROUP BY COALESCE(severity, 'Unspecified')
            """
        ).fetchall()

        touch_rows = conn.execute(
            f"""
            SELECT r.created_at as created_at, MIN(e.created_at) as first_update_at
            FROM risks r
            LEFT JOIN risk_events e
                ON e.risk_id = r.id
               AND e.event_type = 'update'
            {where_sql}
            GROUP BY r.id
            """,
            params,
        ).fetchall()

        closed_rows = conn.execute(
            f"""
            SELECT created_at, closed_at
            FROM risks
            {_where_join(where_sql, "TRIM(COALESCE(closed_at, '')) <> ''")}
            """,
            params,
        ).fetchall()

        # Task queue metrics (global; independent of risk filters)
        today = _today_ymd()
        tasks_overdue = int(
            conn.execute(
                """
                SELECT COUNT(*)
                FROM risk_tasks
                WHERE LOWER(TRIM(status)) NOT IN ('done','cancelled')
                    AND TRIM(COALESCE(due_date,'')) <> ''
                    AND DATE(due_date) < DATE(?)
                """.strip(),
                (today,),
            ).fetchone()[0]
        )
        tasks_due_7 = int(
            conn.execute(
                """
                SELECT COUNT(*)
                FROM risk_tasks
                WHERE LOWER(TRIM(status)) NOT IN ('done','cancelled')
                    AND TRIM(COALESCE(due_date,'')) <> ''
                    AND DATE(due_date) >= DATE(?)
                    AND DATE(due_date) <= DATE(?, '+7 days')
                """.strip(),
                (today, today),
            ).fetchone()[0]
        )

    # Queue counts should be stable and global (not dependent on the current filter view).
    queue_where_sql = " WHERE " + _not_deleted_condition()
    queue_params: list[object] = []

    open_condition = "LOWER(TRIM(COALESCE(status,''))) NOT IN ('closed','archived','mitigated')"

    with get_db_connection() as conn:
        unassigned_where = _where_join(
            queue_where_sql,
            f"TRIM(COALESCE(assigned_to, '')) = '' AND {open_condition}",
        )
        open_count = int(
            conn.execute(
                f"SELECT COUNT(*) FROM risks{_where_join(queue_where_sql, open_condition)}",
                queue_params,
            ).fetchone()[0]
        )
        unassigned_count = int(
            conn.execute(
                f"SELECT COUNT(*) FROM risks{unassigned_where}",
                queue_params,
            ).fetchone()[0]
        )
        high_count = int(
            conn.execute(
                f"""
                SELECT COUNT(*)
                FROM risks
                {_where_join(
                    queue_where_sql,
                    f"LOWER(TRIM(COALESCE(severity,''))) IN ('high','critical') AND {open_condition}",
                )}
                """.strip(),
                queue_params,
            ).fetchone()[0]
        )
        my_count = 0
        me_sql, me_params = _assignee_me_condition(current_user)
        if me_sql:
            my_count = int(
                conn.execute(
                    f"""
                    SELECT COUNT(*)
                    FROM risks
                    {_where_join(queue_where_sql, f"{me_sql} AND {open_condition}")}
                    """.strip(),
                    queue_params + me_params,
                ).fetchone()[0]
            )

    # SLA-style metrics
    first_touch_hours: list[float] = []
    for row in touch_rows:
        created = _parse_date(row["created_at"])
        touched = _parse_date(row["first_update_at"])
        if created and touched and touched >= created:
            first_touch_hours.append((touched - created).total_seconds() / 3600.0)

    closed_days: list[float] = []
    for row in closed_rows:
        created = _parse_date(row["created_at"])
        closed = _parse_date(row["closed_at"])
        if created and closed and closed >= created:
            closed_days.append((closed - created).total_seconds() / (3600.0 * 24.0))

    median_first_touch = _median(first_touch_hours)
    median_close_days = _median(closed_days)

    query_args = {
        "status": status_filter,
        "severity": severity_filter,
        "assigned": assignee_filter,
        "q": query_text,
        "per_page": int(pager["per_page"]),
    }
    if pager.get("has_prev"):
        pager["prev_url"] = url_for("admin_dashboard", **query_args, page=int(pager["page"]) - 1)
    if pager.get("has_next"):
        pager["next_url"] = url_for("admin_dashboard", **query_args, page=int(pager["page"]) + 1)

    return render_template(
        "admin_dashboard.html",
        risks=risks,
        kpi_cards=program_kpis,
        summary=summary,
        severity_summary=severity_summary,
        statuses=sorted(set(statuses)),
        severities=sorted(set(severities)),
        assignees=sorted(set(assignees)),
        status_filter=status_filter,
        severity_filter=severity_filter,
        assignee_filter=assignee_filter,
        query_text=query_text,
        pagination=pager,
        query_args=query_args,
        queue_cards={
            "Open": open_count,
            "High/Critical": high_count,
            "Unassigned": unassigned_count,
            "My queue": my_count,
            "Overdue tasks": tasks_overdue,
            "Tasks due (7d)": tasks_due_7,
        },
        queue_links={
            "Open": url_for("admin_dashboard", status="all", severity="all", assigned="all"),
            "High/Critical": url_for(
                "admin_dashboard", status="all", severity="high", assigned="all"
            ),
            "Unassigned": url_for(
                "admin_dashboard", status="all", severity="all", assigned="unassigned"
            ),
            "My queue": url_for(
                "admin_dashboard", status="all", severity="all", assigned="me"
            ),
            "Overdue tasks": url_for("admin_tasks", due="overdue"),
            "Tasks due (7d)": url_for("admin_tasks", due="due_7"),
        },
        sla_cards={
            "Median first touch": (f"{median_first_touch:.1f}h" if median_first_touch is not None else "—"),
            "Median time to close": (f"{median_close_days:.1f}d" if median_close_days is not None else "—"),
        },
    )


@app.route("/admin/deleted")
def admin_deleted():
    if not require_admin():
        return redirect(url_for("admin"))

    query_text = request.args.get("q", "").strip()
    page, per_page = _parse_pagination(default_per_page=50)

    filters: list[str] = ["TRIM(COALESCE(deleted_at,'')) <> ''"]
    params: list[object] = []

    if query_text:
        filters.append("(LOWER(title) LIKE ? OR LOWER(description) LIKE ?)")
        like = f"%{query_text.lower()}%"
        params.extend([like, like])

    where_sql = " WHERE " + " AND ".join(filters)
    list_query = f"SELECT * FROM risks{where_sql} ORDER BY deleted_at DESC, updated_at DESC LIMIT ? OFFSET ?"

    with get_db_connection() as conn:
        total = int(conn.execute(f"SELECT COUNT(*) FROM risks{where_sql}", params).fetchone()[0])
        pager = _pager(total=total, page=page, per_page=per_page)
        offset = (int(pager["page"]) - 1) * per_page
        risks = conn.execute(list_query, params + [per_page, offset]).fetchall()

    query_args = {
        "q": query_text,
        "per_page": int(pager["per_page"]),
    }
    if pager.get("has_prev"):
        pager["prev_url"] = url_for("admin_deleted", **query_args, page=int(pager["page"]) - 1)
    if pager.get("has_next"):
        pager["next_url"] = url_for("admin_deleted", **query_args, page=int(pager["page"]) + 1)

    return render_template(
        "admin_deleted.html",
        risks=risks,
        query_text=query_text,
        pagination=pager,
        query_args=query_args,
    )


@app.route("/admin/tasks")
def admin_tasks():
    if not require_admin():
        return redirect(url_for("admin"))

    status_filter = request.args.get("status", "open").strip().lower()
    assignee_filter = request.args.get("assigned", "all").strip().lower()
    due_filter = request.args.get("due", "all").strip().lower()
    query_text = request.args.get("q", "").strip()

    current_user = _current_user_email()
    today = _today_ymd()

    filters: list[str] = []
    params: list[object] = []

    # Exclude tasks for soft-deleted risks.
    filters.append(_not_deleted_condition("r"))

    if status_filter == "open":
        filters.append("LOWER(TRIM(t.status)) NOT IN ('done','cancelled')")
    elif status_filter == "done":
        filters.append("LOWER(TRIM(t.status)) = 'done'")
    elif status_filter == "cancelled":
        filters.append("LOWER(TRIM(t.status)) = 'cancelled'")

    if assignee_filter == "me" and current_user:
        filters.append("LOWER(TRIM(COALESCE(t.assigned_to,''))) = ?")
        params.append(current_user.lower())
    elif assignee_filter == "unassigned":
        filters.append("TRIM(COALESCE(t.assigned_to,'')) = ''")

    if due_filter == "overdue":
        filters.append("TRIM(COALESCE(t.due_date,'')) <> '' AND DATE(t.due_date) < DATE(?)")
        params.append(today)
    elif due_filter == "due_7":
        filters.append("TRIM(COALESCE(t.due_date,'')) <> '' AND DATE(t.due_date) <= DATE(?, '+7 days')")
        params.append(today)
    elif due_filter == "due_30":
        filters.append("TRIM(COALESCE(t.due_date,'')) <> '' AND DATE(t.due_date) <= DATE(?, '+30 days')")
        params.append(today)
    elif due_filter == "no_due":
        filters.append("TRIM(COALESCE(t.due_date,'')) = ''")

    if query_text:
        filters.append("(LOWER(t.title) LIKE ? OR LOWER(COALESCE(t.notes,'')) LIKE ? OR LOWER(r.title) LIKE ?)")
        like = f"%{query_text.lower()}%"
        params.extend([like, like, like])

    where_sql = (" WHERE " + " AND ".join(filters)) if filters else ""

    with get_db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT
                t.*, r.title as risk_title, r.severity as risk_severity, r.status as risk_status
            FROM risk_tasks t
            JOIN risks r ON r.id = t.risk_id
            {where_sql}
            ORDER BY
                CASE WHEN LOWER(TRIM(t.status)) IN ('done','cancelled') THEN 1 ELSE 0 END,
                CASE WHEN TRIM(COALESCE(t.due_date,'')) = '' THEN 1 ELSE 0 END,
                t.due_date ASC,
                t.id DESC
            """.strip(),
            params,
        ).fetchall()

        assignees = [
            row["assigned_to"]
            for row in conn.execute(
                "SELECT DISTINCT TRIM(assigned_to) as assigned_to FROM risk_tasks WHERE assigned_to IS NOT NULL AND TRIM(assigned_to) <> ''"
            )
        ]

    return render_template(
        "admin_tasks.html",
        tasks=rows,
        task_status_options=TASK_STATUS_OPTIONS,
        assignees=sorted(set(assignees)),
        status_filter=status_filter,
        assignee_filter=assignee_filter,
        due_filter=due_filter,
        query_text=query_text,
        today=today,
    )


@app.route("/admin/heatmap")
def admin_heatmap():
    if not require_admin():
        return redirect(url_for("admin"))

    scope = (request.args.get("scope") or "open").strip().lower()
    if scope not in {"open", "all"}:
        scope = "open"

    levels = ["Very Low", "Low", "Medium", "High", "Very High"]

    def to_level(v: str | None) -> int | None:
        return _rating_1_to_5(v)

    def bump(matrix: dict[tuple[int, int], int], like: int | None, imp: int | None) -> None:
        if like is None or imp is None:
            return
        matrix[(like, imp)] = matrix.get((like, imp), 0) + 1

    initial_matrix: dict[tuple[int, int], int] = {}
    residual_matrix: dict[tuple[int, int], int] = {}
    initial_unscored = 0
    residual_unscored = 0

    band_counts_initial: dict[str, int] = {"Low": 0, "Medium": 0, "High": 0, "Critical": 0, "Unscored": 0}
    band_counts_residual: dict[str, int] = {"Low": 0, "Medium": 0, "High": 0, "Critical": 0, "Unscored": 0}

    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT COUNT(*) as c FROM risks WHERE TRIM(COALESCE(deleted_at,'')) <> ''"
        ).fetchone()
        deleted_excluded = int(row["c"] if row else 0)

        row = conn.execute(
            "SELECT COUNT(*) as c FROM risks WHERE TRIM(COALESCE(deleted_at,'')) = ''"
        ).fetchone()
        not_deleted_total = int(row["c"] if row else 0)

        where = ["TRIM(COALESCE(deleted_at,'')) = ''"]
        if scope == "open":
            where.append("LOWER(TRIM(COALESCE(status,''))) NOT IN ('closed','archived','mitigated')")

        risks = conn.execute(
            """
            SELECT id, title, status, likelihood_initial, impact_initial, likelihood_residual, impact_residual, likelihood, impact
            FROM risks
            WHERE {where_sql}
            ORDER BY updated_at DESC
            """.strip().format(where_sql=" AND ".join(where)),
        ).fetchall()

    included_total = len(risks)
    nonopen_excluded = max(0, not_deleted_total - included_total) if scope == "open" else 0

    for r in risks:
        li = to_level(r["likelihood_initial"] or r["likelihood"])
        ii = to_level(r["impact_initial"] or r["impact"])
        lr = to_level(r["likelihood_residual"])
        ir = to_level(r["impact_residual"])

        if li is None or ii is None:
            initial_unscored += 1
            band_counts_initial["Unscored"] += 1
        else:
            bump(initial_matrix, li, ii)
            band_counts_initial[_score_band(_score_1_to_25(li, ii))] += 1

        if lr is None or ir is None:
            residual_unscored += 1
            band_counts_residual["Unscored"] += 1
        else:
            bump(residual_matrix, lr, ir)
            band_counts_residual[_score_band(_score_1_to_25(lr, ir))] += 1

    return render_template(
        "admin_heatmap.html",
        levels=levels,
        initial_matrix=initial_matrix,
        residual_matrix=residual_matrix,
        initial_unscored=initial_unscored,
        residual_unscored=residual_unscored,
        band_counts_initial=band_counts_initial,
        band_counts_residual=band_counts_residual,
        scope=scope,
        included_total=included_total,
        not_deleted_total=not_deleted_total,
        deleted_excluded=deleted_excluded,
        nonopen_excluded=nonopen_excluded,
    )


@app.route("/admin/trends")
def admin_trends():
    if not require_admin():
        return redirect(url_for("admin"))

    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                snapshot_date,
                COUNT(*) as total,
                SUM(CASE WHEN level_initial IN ('High','Critical') THEN 1 ELSE 0 END) as hi_init,
                SUM(CASE WHEN level_residual IN ('High','Critical') THEN 1 ELSE 0 END) as hi_res,
                SUM(CASE WHEN level_initial = 'Critical' THEN 1 ELSE 0 END) as crit_init,
                SUM(CASE WHEN level_residual = 'Critical' THEN 1 ELSE 0 END) as crit_res
            FROM risk_score_snapshots
            GROUP BY snapshot_date
            ORDER BY snapshot_date DESC
            LIMIT 60
            """.strip()
        ).fetchall()

        latest = conn.execute(
            "SELECT MAX(snapshot_date) as latest_date FROM risk_score_snapshots"
        ).fetchone()
        latest_date = (latest["latest_date"] if latest else None) or ""

    return render_template(
        "admin_trends.html",
        rows=rows,
        latest_date=latest_date,
        today=_today_ymd(),
    )


@app.route("/admin/trends/snapshot", methods=["POST"])
def admin_trends_snapshot():
    if not require_admin():
        return redirect(url_for("admin"))

    actor = _current_actor()
    now = datetime.utcnow().isoformat()
    snap_date = _today_ymd()

    inserted = 0
    with get_db_connection() as conn:
        risks = conn.execute(
            """
            SELECT id, likelihood_initial, impact_initial, likelihood_residual, impact_residual, likelihood, impact
            FROM risks
            WHERE TRIM(COALESCE(deleted_at,'')) = ''
            """.strip()
        ).fetchall()

        for r in risks:
            rid = int(r["id"])
            li = _rating_1_to_5(r["likelihood_initial"] or r["likelihood"])
            ii = _rating_1_to_5(r["impact_initial"] or r["impact"])
            lr = _rating_1_to_5(r["likelihood_residual"])
            ir = _rating_1_to_5(r["impact_residual"])

            score_i = _score_1_to_25(li, ii)
            score_r = _score_1_to_25(lr, ir)
            if score_i is None and score_r is None:
                continue

            cur = conn.execute(
                """
                INSERT OR REPLACE INTO risk_score_snapshots
                    (snapshot_date, risk_id, score_initial, level_initial, score_residual, level_residual, created_by, created_at)
                VALUES
                    (?, ?, ?, ?, ?, ?, ?, ?)
                """.strip(),
                (
                    snap_date,
                    rid,
                    score_i,
                    _score_band(score_i),
                    score_r,
                    _score_band(score_r),
                    actor,
                    now,
                ),
            )
            if int(cur.rowcount or 0) > 0:
                inserted += 1

        conn.commit()

    flash(f"Snapshot captured for {inserted} risk(s) on {snap_date}.", "success")
    return redirect(url_for("admin_trends"))


def _send_task_reminders(*, days_ahead: int = 3) -> tuple[int, int]:
    """Send reminder emails for tasks due within the next N days (including overdue).

    Returns: (recipient_count, task_count)
    """

    days_ahead = max(0, min(int(days_ahead), 60))
    cutoff = datetime.utcnow().date().isoformat()

    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                t.id as task_id,
                t.risk_id as risk_id,
                t.title as task_title,
                t.due_date as due_date,
                t.assigned_to as assigned_to,
                r.title as risk_title
            FROM risk_tasks t
            JOIN risks r ON r.id = t.risk_id
            WHERE LOWER(TRIM(t.status)) NOT IN ('done','cancelled')
                            AND TRIM(COALESCE(r.deleted_at,'')) = ''
              AND TRIM(COALESCE(t.assigned_to,'')) <> ''
              AND TRIM(COALESCE(t.due_date,'')) <> ''
              AND DATE(t.due_date) <= DATE(?, '+' || ? || ' days')
            ORDER BY LOWER(TRIM(t.assigned_to)), DATE(t.due_date), t.id
            """.strip(),
            (cutoff, str(days_ahead)),
        ).fetchall()

        by_recipient: dict[str, list[sqlite3.Row]] = {}
        for row in rows:
            addr = (row["assigned_to"] or "").strip()
            if not addr or "@" not in addr:
                continue
            by_recipient.setdefault(addr, []).append(row)

        base_url = os.getenv("APP_BASE_URL", "").strip() or request.url_root.rstrip("/")
        actor = _current_actor()
        now = datetime.utcnow().isoformat()
        recipients_sent = 0
        tasks_in_emails = 0

        for addr, items in by_recipient.items():
            lines: list[str] = []
            for it in items:
                link = f"{base_url}{url_for('admin_risk_detail', risk_id=int(it['risk_id']))}"
                due = (it["due_date"] or "")[:10]
                lines.append(
                    f"<li><strong>{(it['risk_title'] or '').strip()}</strong>: {it['task_title']} (due {due}) — <a href=\"{link}\">open</a></li>"
                )

            if not lines:
                continue

            html = (
                f"<p>You have {len(lines)} risk task(s) due within {days_ahead} day(s) (including overdue).</p>"
                f"<ul>{''.join(lines)}</ul>"
            )

            sent = _send_email(
                subject=f"Risk Tasks Due (next {days_ahead}d)",
                html_body=html,
                to_addrs=[addr],
            )

            if sent:
                recipients_sent += 1
                tasks_in_emails += len(lines)
                for it in items:
                    _log_event(
                        conn,
                        risk_id=int(it["risk_id"]),
                        event_type="notify",
                        field="task_reminder",
                        old_value="",
                        new_value=addr,
                        actor=actor,
                    )

        conn.commit()

    return recipients_sent, tasks_in_emails


@app.route("/admin/tasks/reminders/send", methods=["POST"])
def admin_send_task_reminders():
    if not require_admin():
        return redirect(url_for("admin"))

    try:
        days_ahead = int(float((request.form.get("days_ahead") or "").strip() or "3"))
    except ValueError:
        days_ahead = 3

    recipients_sent, tasks_in_emails = _send_task_reminders(days_ahead=days_ahead)
    flash(f"Sent {recipients_sent} reminder email(s) covering {tasks_in_emails} task(s).", "success")
    return redirect(request.referrer or url_for("admin_tasks"))


@app.route("/admin/risk/<int:risk_id>/tasks/add", methods=["POST"])
def admin_task_add(risk_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    title = (request.form.get("title") or "").strip()
    assigned_to = (request.form.get("assigned_to") or "").strip()
    due_date = _parse_ymd(request.form.get("due_date"))
    notes = (request.form.get("notes") or "").strip()

    if not title:
        flash("Task title is required.", "warning")
        return redirect(url_for("admin_risk_detail", risk_id=risk_id))

    if assigned_to:
        ok, normalized = _validate_assignee_email(assigned_to)
        if not ok:
            allowlist, _src = _allowed_assignee_emails()
            if normalized and allowlist:
                flash("Task assignee must be a member of the Risk Management group.", "warning")
            else:
                flash("Task assignee must be a valid email address (e.g. name@hdh.org).", "warning")
            return redirect(url_for("admin_risk_detail", risk_id=risk_id))
        assigned_to = normalized

    actor = _current_actor()
    now = datetime.utcnow().isoformat()

    with get_db_connection() as conn:
        exists = conn.execute("SELECT 1 FROM risks WHERE id = ?", (risk_id,)).fetchone()
        if not exists:
            flash("Risk not found.", "warning")
            return redirect(url_for("admin_dashboard"))

        cur = conn.execute(
            """
            INSERT INTO risk_tasks (risk_id, title, notes, status, assigned_to, due_date, created_by, updated_by, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """.strip(),
            (
                risk_id,
                title,
                notes,
                "Open",
                assigned_to,
                due_date,
                actor,
                actor,
                now,
                now,
            ),
        )
        task_id = int(cur.lastrowid)
        _log_event(
            conn,
            risk_id=risk_id,
            event_type="task_create",
            field="task",
            old_value="",
            new_value=f"#{task_id} {title}"[:500],
            actor=actor,
        )
        conn.execute(
            "UPDATE risks SET updated_by = ?, updated_at = ? WHERE id = ?",
            (actor, now, risk_id),
        )
        conn.commit()

    flash("Task added.", "success")
    return redirect(url_for("admin_risk_detail", risk_id=risk_id))


@app.route("/admin/tasks/<int:task_id>/status", methods=["POST"])
def admin_task_set_status(task_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    status = _normalize_task_status(request.form.get("status"))
    actor = _current_actor()
    now = datetime.utcnow().isoformat()

    with get_db_connection() as conn:
        task = conn.execute("SELECT * FROM risk_tasks WHERE id = ?", (task_id,)).fetchone()
        if task is None:
            flash("Task not found.", "warning")
            return redirect(request.referrer or url_for("admin_tasks"))

        risk_id = int(task["risk_id"])
        old_status = (task["status"] or "").strip()

        completed_at = (task["completed_at"] or "").strip()
        if status == "Done" and not completed_at:
            completed_at = now
        if status != "Done":
            completed_at = ""

        conn.execute(
            """
            UPDATE risk_tasks
            SET status = ?, completed_at = ?, updated_by = ?, updated_at = ?
            WHERE id = ?
            """.strip(),
            (status, completed_at, actor, now, task_id),
        )
        if old_status != status:
            _log_event(
                conn,
                risk_id=risk_id,
                event_type="task_update",
                field="task_status",
                old_value=f"#{task_id} {old_status}"[:500],
                new_value=f"#{task_id} {status}"[:500],
                actor=actor,
            )
        conn.execute(
            "UPDATE risks SET updated_by = ?, updated_at = ? WHERE id = ?",
            (actor, now, risk_id),
        )
        conn.commit()

    flash("Task updated.", "success")
    return redirect(request.referrer or url_for("admin_tasks"))


@app.route("/admin/tasks/<int:task_id>/delete", methods=["POST"])
def admin_task_delete(task_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    actor = _current_actor()
    now = datetime.utcnow().isoformat()

    with get_db_connection() as conn:
        task = conn.execute("SELECT * FROM risk_tasks WHERE id = ?", (task_id,)).fetchone()
        if task is None:
            flash("Task not found.", "warning")
            return redirect(request.referrer or url_for("admin_tasks"))
        risk_id = int(task["risk_id"])
        title = (task["title"] or "").strip()
        conn.execute("DELETE FROM risk_tasks WHERE id = ?", (task_id,))
        _log_event(
            conn,
            risk_id=risk_id,
            event_type="task_delete",
            field="task",
            old_value=f"#{task_id} {title}"[:500],
            new_value="",
            actor=actor,
        )
        conn.execute(
            "UPDATE risks SET updated_by = ?, updated_at = ? WHERE id = ?",
            (actor, now, risk_id),
        )
        conn.commit()

    flash("Task deleted.", "success")
    return redirect(request.referrer or url_for("admin_tasks"))


@app.route("/admin/risk/<int:risk_id>/watchers/add", methods=["POST"])
def admin_risk_watchers_add(risk_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    emails_raw = (request.form.get("email") or request.form.get("emails") or "").strip()
    emails = [_normalize_email(e) for e in _parse_recipients(emails_raw)]
    emails = [e for e in emails if e]
    if not emails:
        flash("Enter at least one valid email address.", "warning")
        return redirect(url_for("admin_risk_detail", risk_id=risk_id) + "#comments")

    actor = _current_actor()
    now = datetime.utcnow().isoformat()
    added = 0

    with get_db_connection() as conn:
        exists = conn.execute("SELECT 1 FROM risks WHERE id = ?", (risk_id,)).fetchone()
        if not exists:
            flash("Risk not found.", "warning")
            return redirect(url_for("admin_dashboard"))

        for email in emails:
            cur = conn.execute(
                """
                INSERT OR IGNORE INTO risk_watchers (risk_id, email, created_by, created_at)
                VALUES (?, ?, ?, ?)
                """.strip(),
                (risk_id, email, actor, now),
            )
            if int(cur.rowcount or 0) > 0:
                added += 1

        if added:
            _log_event(
                conn,
                risk_id=risk_id,
                event_type="watchers_add",
                field="watchers",
                old_value="",
                new_value=",".join(emails)[:500],
                actor=actor,
            )

        conn.execute(
            "UPDATE risks SET updated_by = ?, updated_at = ? WHERE id = ?",
            (actor, now, risk_id),
        )
        conn.commit()

    if added:
        flash(f"Added {added} watcher(s).", "success")
    else:
        flash("No new watchers were added (already subscribed).", "info")
    return redirect(url_for("admin_risk_detail", risk_id=risk_id) + "#comments")


@app.route("/admin/risk/<int:risk_id>/watchers/<int:watcher_id>/remove", methods=["POST"])
def admin_risk_watchers_remove(risk_id: int, watcher_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    actor = _current_actor()
    now = datetime.utcnow().isoformat()

    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM risk_watchers WHERE id = ? AND risk_id = ?",
            (watcher_id, risk_id),
        ).fetchone()
        if row is None:
            flash("Watcher not found.", "warning")
            return redirect(url_for("admin_risk_detail", risk_id=risk_id) + "#comments")

        email = (row["email"] or "").strip()
        conn.execute("DELETE FROM risk_watchers WHERE id = ?", (watcher_id,))
        _log_event(
            conn,
            risk_id=risk_id,
            event_type="watchers_remove",
            field="watchers",
            old_value=email[:500],
            new_value="",
            actor=actor,
        )
        conn.execute(
            "UPDATE risks SET updated_by = ?, updated_at = ? WHERE id = ?",
            (actor, now, risk_id),
        )
        conn.commit()

    flash("Watcher removed.", "success")
    return redirect(url_for("admin_risk_detail", risk_id=risk_id) + "#comments")


@app.route("/admin/risk/<int:risk_id>/comments/add", methods=["POST"])
def admin_risk_comment_add(risk_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    body = (request.form.get("body") or "").strip()
    if not body:
        flash("Comment cannot be empty.", "warning")
        return redirect(url_for("admin_risk_detail", risk_id=risk_id) + "#comments")

    actor = _current_actor()
    now = datetime.utcnow().isoformat()

    with get_db_connection() as conn:
        risk = conn.execute("SELECT * FROM risks WHERE id = ?", (risk_id,)).fetchone()
        if risk is None:
            flash("Risk not found.", "warning")
            return redirect(url_for("admin_dashboard"))

        is_deleted = bool((risk["deleted_at"] or "").strip())
        if request.method == "POST" and is_deleted:
            flash("This risk is deleted. Restore it before editing.", "warning")
            return redirect(url_for("admin_risk_detail", risk_id=risk_id))

        cur = conn.execute(
            """
            INSERT INTO risk_comments (risk_id, body, author, created_at)
            VALUES (?, ?, ?, ?)
            """.strip(),
            (risk_id, body, actor, now),
        )
        comment_id = int(cur.lastrowid)

        _log_event(
            conn,
            risk_id=risk_id,
            event_type="comment",
            field="comment",
            old_value="",
            new_value=(body.replace("\n", " ")[:500]),
            actor=actor,
        )

        conn.execute(
            "UPDATE risks SET updated_by = ?, updated_at = ? WHERE id = ?",
            (actor, now, risk_id),
        )

        # Notify watchers
        if _env_bool("NOTIFY_COMMENTS_ENABLED", True):
            watcher_rows = conn.execute(
                "SELECT email FROM risk_watchers WHERE risk_id = ? ORDER BY LOWER(email)",
                (risk_id,),
            ).fetchall()
            to_addrs = []
            actor_email = _normalize_email(actor)
            for wr in watcher_rows:
                addr = _normalize_email(wr["email"])
                if not addr:
                    continue
                if actor_email and addr == actor_email:
                    continue
                to_addrs.append(addr)

            if to_addrs:
                base_url = os.getenv("APP_BASE_URL", "").strip() or request.url_root.rstrip("/")
                link = f"{base_url}{url_for('admin_risk_detail', risk_id=risk_id)}#comments"
                title = (risk["title"] or "Risk")
                html = (
                    f"<p><strong>New comment</strong> on risk <strong>{title}</strong></p>"
                    f"<p><strong>Author:</strong> {actor}</p>"
                    f"<p style=\"white-space:pre-wrap\">{body}</p>"
                    f"<p><a href=\"{link}\">Open risk</a></p>"
                )
                sent = _send_email(
                    subject=f"Risk Comment: {title}",
                    html_body=html,
                    to_addrs=sorted(set(to_addrs)),
                )
                if sent:
                    _log_event(
                        conn,
                        risk_id=risk_id,
                        event_type="notify",
                        field="comment_watchers",
                        old_value=str(comment_id),
                        new_value=",".join(sorted(set(to_addrs)))[:500],
                        actor=actor,
                    )

        conn.commit()

    flash("Comment added.", "success")
    return redirect(url_for("admin_risk_detail", risk_id=risk_id) + "#comments")


@app.route("/admin/bulk_update", methods=["POST"])
def admin_bulk_update():
    if not require_admin():
        return redirect(url_for("admin"))

    selected_ids = request.form.getlist("risk_id")
    status = (request.form.get("bulk_status") or "").strip()
    severity = (request.form.get("bulk_severity") or "").strip()
    assigned_to = (request.form.get("bulk_assigned_to") or "").strip()
    clear_assigned_to = (request.form.get("bulk_clear_assigned_to") or "").strip() in {"1", "true", "on", "yes"}
    close_reason = (request.form.get("bulk_close_reason") or "").strip()
    return_to = (request.form.get("return_to") or "").strip()

    # Treat blank fields as "no change".
    fields_to_update: dict[str, str] = {}
    if status:
        fields_to_update["status"] = status
    if severity:
        if severity not in SEVERITY_OPTIONS:
            flash("Invalid severity selection.", "warning")
            return redirect(return_to or url_for("admin_dashboard"))
        fields_to_update["severity"] = severity
    if clear_assigned_to:
        fields_to_update["assigned_to"] = ""
    elif assigned_to:
        ok, normalized = _validate_assignee_email(assigned_to)
        if not ok:
            allowlist, _src = _allowed_assignee_emails()
            if normalized and allowlist:
                flash("Assignee must be a member of the Risk Management group.", "warning")
            else:
                flash("Assignee must be a valid email address (e.g. name@hdh.org).", "warning")
            return redirect(return_to or url_for("admin_dashboard"))
        fields_to_update["assigned_to"] = normalized
    if close_reason:
        fields_to_update["close_reason"] = close_reason

    if status == "Closed" and not close_reason:
        flash("Close reason is required when bulk-closing risks.", "warning")
        return redirect(return_to or url_for("admin_dashboard"))

    if not selected_ids:
        flash("Select at least one risk.", "warning")
        return redirect(return_to or url_for("admin_dashboard"))
    if not fields_to_update:
        flash("Choose at least one bulk action (status, severity, or assignee).", "warning")
        return redirect(return_to or url_for("admin_dashboard"))

    try:
        risk_ids = [int(x) for x in selected_ids]
    except ValueError:
        flash("Invalid selection.", "danger")
        return redirect(return_to or url_for("admin_dashboard"))

    actor = _current_actor()
    now = datetime.utcnow().isoformat()
    base_url = os.getenv("APP_BASE_URL", "").strip() or request.url_root.rstrip("/")
    updated_count = 0
    with get_db_connection() as conn:
        existing = conn.execute(
            f"SELECT * FROM risks WHERE id IN ({','.join('?' for _ in risk_ids)}) AND {_not_deleted_condition()}",
            risk_ids,
        ).fetchall()
        existing_by_id = {int(r["id"]): r for r in existing}

        for rid in risk_ids:
            risk = existing_by_id.get(rid)
            if risk is None:
                continue

            old_assignee = _normalize_email(risk["assigned_to"])

            for field, new_value in fields_to_update.items():
                old_value = (risk[field] or "") if field in risk.keys() else ""
                if str(old_value).strip() != str(new_value).strip():
                    _log_event(
                        conn,
                        risk_id=rid,
                        event_type="update",
                        field=field,
                        old_value=str(old_value),
                        new_value=str(new_value),
                        actor=actor,
                    )

            # closed_at/close_reason rules only apply when setting Closed via bulk
            closed_at = (risk["closed_at"] or "").strip()
            if fields_to_update.get("status") == "Closed" and not closed_at:
                closed_at = now
                _log_event(
                    conn,
                    risk_id=rid,
                    event_type="update",
                    field="closed_at",
                    old_value="",
                    new_value=closed_at,
                    actor=actor,
                )

            sets = []
            params: list[object] = []
            for field, value in fields_to_update.items():
                sets.append(f"{field} = ?")
                params.append(value)

            sets.append("updated_by = ?")
            params.append(actor)

            if fields_to_update.get("status") == "Closed":
                sets.append("closed_at = ?")
                params.append(closed_at)

            sets.append("updated_at = ?")
            params.append(now)
            params.append(rid)

            conn.execute(
                f"UPDATE risks SET {', '.join(sets)} WHERE id = ? AND {_not_deleted_condition()}",
                params,
            )
            updated_count += 1

            # Notifications: assignment change
            new_assignee = (fields_to_update.get("assigned_to") if "assigned_to" in fields_to_update else old_assignee) or ""
            new_assignee = _normalize_email(str(new_assignee))
            if new_assignee and new_assignee != old_assignee:
                try:
                    if _env_bool("NOTIFY_ASSIGNMENT_ENABLED", True):
                        link = f"{base_url}{url_for('admin_risk_detail', risk_id=rid)}"
                        owner_email = _normalize_email(risk["owner"] if "owner" in risk.keys() else "")
                        to_addrs = _unique_emails([new_assignee, owner_email])
                        if not to_addrs:
                            continue
                        status_for_email = (fields_to_update.get("status") or risk["status"] or "").strip()
                        severity_for_email = (fields_to_update.get("severity") or risk["severity"] or "").strip()
                        sent = _send_email(
                            subject=f"Risk Assigned: {risk['title']}",
                            html_body=(
                                f"<p>You have been assigned a risk.</p>"
                                f"<p><strong>Title:</strong> {risk['title']}</p>"
                                f"<p><strong>Status:</strong> {status_for_email}</p>"
                                f"<p><strong>Severity:</strong> {severity_for_email or 'Unspecified'}</p>"
                                f"<p><a href=\"{link}\">Open risk</a></p>"
                            ),
                            to_addrs=to_addrs,
                        )
                        if sent:
                            _log_event(
                                conn,
                                risk_id=rid,
                                event_type="notify",
                                field="email",
                                old_value="",
                                new_value=",".join(to_addrs)[:500],
                                actor=actor,
                            )
                except Exception:
                    pass

        conn.commit()

    flash(f"Bulk updated {updated_count} risk(s).", "success")
    return redirect(return_to or url_for("admin_dashboard"))


@app.route("/admin/export.csv")
def admin_export_csv():
    if not require_admin():
        return redirect(url_for("admin"))

    status_filter = request.args.get("status", "all").strip()
    severity_filter = request.args.get("severity", "all").strip()
    assignee_filter = request.args.get("assigned", "all").strip()
    owner_filter = request.args.get("owner", "all").strip()
    query_text = request.args.get("q", "").strip()
    date_field = request.args.get("date_field", "created_at").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    current_user = _current_user_email()
    where_sql, params = _build_admin_export_where(
        status_filter=status_filter,
        severity_filter=severity_filter,
        assignee_filter=assignee_filter,
        owner_filter=owner_filter,
        query_text=query_text,
        current_user=current_user,
        date_field=date_field,
        date_from=date_from,
        date_to=date_to,
    )
    query = f"SELECT * FROM risks{where_sql} ORDER BY updated_at DESC"

    with get_db_connection() as conn:
        risks = conn.execute(query, params).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "id",
            "title",
            "severity",
            "status",
            "assigned_to",
            "owner",
            "priority",
            "progress",
            "date_identified",
            "impact_type",
            "created_at",
            "updated_at",
            "closed_at",
            "close_reason",
        ]
    )
    for r in risks:
        writer.writerow(
            [
                r["id"],
                r["title"],
                r["severity"],
                r["status"],
                r["assigned_to"],
                r["owner"],
                r["priority"],
                r["progress"],
                r["date_identified"],
                r["impact_type"],
                r["created_at"],
                r["updated_at"],
                r["closed_at"],
                r["close_reason"],
            ]
        )

    csv_bytes = output.getvalue().encode("utf-8")
    response = app.response_class(csv_bytes, mimetype="text/csv; charset=utf-8")
    response.headers["Content-Disposition"] = "attachment; filename=risks_export.csv"
    return response


@app.route("/admin/export")
def admin_export():
    if not require_admin():
        return redirect(url_for("admin"))

    status_filter = request.args.get("status", "all").strip()
    severity_filter = request.args.get("severity", "all").strip()
    assignee_filter = request.args.get("assigned", "all").strip()
    owner_filter = request.args.get("owner", "all").strip()
    query_text = request.args.get("q", "").strip()
    date_field = request.args.get("date_field", "created_at").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    current_user = _current_user_email()
    where_sql, where_params = _build_admin_export_where(
        status_filter=status_filter,
        severity_filter=severity_filter,
        assignee_filter=assignee_filter,
        owner_filter=owner_filter,
        query_text=query_text,
        current_user=current_user,
        date_field=date_field,
        date_from=date_from,
        date_to=date_to,
    )

    with get_db_connection() as conn:
        match_count = int(conn.execute(f"SELECT COUNT(*) FROM risks{where_sql}", where_params).fetchone()[0])
        statuses = [
            row["status"]
            for row in conn.execute(
                "SELECT DISTINCT TRIM(status) as status FROM risks WHERE status IS NOT NULL AND TRIM(status) <> ''"
            )
        ]
        severities = [
            row["severity"]
            for row in conn.execute(
                "SELECT DISTINCT TRIM(severity) as severity FROM risks WHERE severity IS NOT NULL AND TRIM(severity) <> ''"
            )
        ]
        assignees = [
            row["assigned_to"]
            for row in conn.execute(
                "SELECT DISTINCT TRIM(assigned_to) as assigned_to FROM risks WHERE assigned_to IS NOT NULL AND TRIM(assigned_to) <> ''"
            )
        ]
        owners = [
            row["owner"]
            for row in conn.execute(
                "SELECT DISTINCT TRIM(owner) as owner FROM risks WHERE owner IS NOT NULL AND TRIM(owner) <> ''"
            )
        ]

    date_fields = [
        ("created_at", "Created"),
        ("updated_at", "Last Updated"),
        ("closed_at", "Closed"),
        ("date_identified", "Date Identified"),
    ]

    return render_template(
        "admin_export.html",
        match_count=match_count,
        statuses=sorted(set(statuses)),
        severities=sorted(set(severities)),
        assignees=sorted(set(assignees)),
        owners=sorted(set(owners)),
        date_fields=date_fields,
        status_filter=status_filter,
        severity_filter=severity_filter,
        assignee_filter=assignee_filter,
        owner_filter=owner_filter,
        query_text=query_text,
        date_field=date_field,
        date_from=date_from,
        date_to=date_to,
    )


@app.route("/admin/import", methods=["POST"])
def admin_import():
    if not require_admin():
        return redirect(url_for("admin"))

    if not app.config.get("ENABLE_ADMIN_IMPORT", False):
        abort(404)

    with get_db_connection() as conn:
        conn.execute("DELETE FROM risks WHERE source_sheet IS NOT NULL")
        conn.execute(
            """
            DELETE FROM risks
            WHERE title = 'Imported Risk'
               OR description = 'Imported from register.'
            """
        )
        conn.execute(
            """
            DELETE FROM risks
            WHERE status = 'New'
              AND (owner IS NULL OR TRIM(owner) = '')
              AND (priority IS NULL OR priority = 0)
              AND progress = 0
              AND (mitigation IS NULL OR TRIM(mitigation) = '')
              AND (impact_type IS NULL OR TRIM(impact_type) = '')
              AND (date_identified IS NULL OR TRIM(date_identified) = '')
            """
        )
        conn.execute("DELETE FROM kpis")
        conn.commit()

    imported = seed_from_excel(force=True)
    flash(f"Imported {imported} risks from the register.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/risk/<int:risk_id>", methods=["GET", "POST"])
def admin_risk_detail(risk_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    try:
        events_page = int(float(request.args.get("events_page", "1")))
    except ValueError:
        events_page = 1
    events_page = max(1, events_page)
    events_per_page = 25

    with get_db_connection() as conn:
        risk = conn.execute("SELECT * FROM risks WHERE id = ?", (risk_id,)).fetchone()
        if risk is None:
            flash("Risk not found.", "warning")
            return redirect(url_for("admin_dashboard"))

        events_total = int(
            conn.execute(
                "SELECT COUNT(*) FROM risk_events WHERE risk_id = ?",
                (risk_id,),
            ).fetchone()[0]
        )
        events_pagination = _pager(total=events_total, page=events_page, per_page=events_per_page)
        events_offset = (int(events_pagination["page"]) - 1) * events_per_page

        events = conn.execute(
            "SELECT * FROM risk_events WHERE risk_id = ? ORDER BY id DESC LIMIT ? OFFSET ?",
            (risk_id, events_per_page, events_offset),
        ).fetchall()

        events_query_args = request.args.to_dict(flat=True)
        if events_pagination["has_prev"]:
            prev_args = dict(events_query_args)
            prev_args["events_page"] = int(events_pagination["page"]) - 1
            events_pagination["prev_url"] = url_for("admin_risk_detail", risk_id=risk_id, **prev_args)
        if events_pagination["has_next"]:
            next_args = dict(events_query_args)
            next_args["events_page"] = int(events_pagination["page"]) + 1
            events_pagination["next_url"] = url_for("admin_risk_detail", risk_id=risk_id, **next_args)

        attachments = conn.execute(
            "SELECT * FROM risk_attachments WHERE risk_id = ? ORDER BY id DESC",
            (risk_id,),
        ).fetchall()

        tasks = conn.execute(
            """
            SELECT *
            FROM risk_tasks
            WHERE risk_id = ?
            ORDER BY
                CASE WHEN LOWER(TRIM(status)) IN ('done','cancelled') THEN 1 ELSE 0 END,
                CASE WHEN TRIM(COALESCE(due_date,'')) = '' THEN 1 ELSE 0 END,
                due_date ASC,
                id DESC
            """.strip(),
            (risk_id,),
        ).fetchall()

        watchers = conn.execute(
            """
            SELECT *
            FROM risk_watchers
            WHERE risk_id = ?
            ORDER BY LOWER(email) ASC
            """.strip(),
            (risk_id,),
        ).fetchall()

        comments = conn.execute(
            """
            SELECT *
            FROM risk_comments
            WHERE risk_id = ?
            ORDER BY id DESC
            """.strip(),
            (risk_id,),
        ).fetchall()

        assignees = [
            row["assigned_to"]
            for row in conn.execute(
                "SELECT DISTINCT TRIM(assigned_to) as assigned_to FROM risks WHERE assigned_to IS NOT NULL AND TRIM(assigned_to) <> ''"
            )
        ]
        allowed_assignees, _assignee_source = _allowed_assignee_emails()
        if allowed_assignees:
            assignees = allowed_assignees
        owners = [
            row["owner"]
            for row in conn.execute(
                "SELECT DISTINCT TRIM(owner) as owner FROM risks WHERE owner IS NOT NULL AND TRIM(owner) <> ''"
            )
        ]

        if request.method == "POST":
            status = request.form.get("status", "New")
            owner = request.form.get("owner", "")
            severity = request.form.get("severity", "")
            assigned_to = request.form.get("assigned_to", "")
            mitigation = request.form.get("mitigation", "")
            admin_notes = request.form.get("admin_notes", "")
            close_reason = request.form.get("close_reason", "")
            progress = request.form.get("progress", "0")
            priority_text = request.form.get("priority", "")
            impact_type = request.form.get("impact_type", "")
            likelihood_initial = request.form.get("likelihood_initial", "")
            impact_initial = request.form.get("impact_initial", "")
            likelihood_residual = request.form.get("likelihood_residual", "")
            impact_residual = request.form.get("impact_residual", "")
            date_identified = request.form.get("date_identified", "")
            review_period = request.form.get("review_period", "")
            date_last_reviewed = request.form.get("date_last_reviewed", "")
            next_review = request.form.get("next_review", "")
            date_postponed = request.form.get("date_postponed", "")
            reason_postponed = request.form.get("reason_postponed", "")
            date_mitigated = request.form.get("date_mitigated", "")
            mitigation_history = request.form.get("mitigation_history", "")

            try:
                progress_value = int(float(progress))
            except ValueError:
                progress_value = 0

            try:
                priority = int(float(priority_text)) if priority_text else None
            except ValueError:
                priority = None

            if not likelihood_residual and likelihood_initial:
                likelihood_residual = likelihood_initial
            if not impact_residual and impact_initial:
                impact_residual = impact_initial
            if not likelihood_initial and likelihood_residual:
                likelihood_initial = likelihood_residual
            if not impact_initial and impact_residual:
                impact_initial = impact_residual

            likelihood_value = likelihood_residual or likelihood_initial
            impact_value = impact_residual or impact_initial

            severity = (severity or "").strip()
            if not severity:
                severity = _compute_severity(likelihood_value, impact_value)
            if severity not in SEVERITY_OPTIONS:
                severity = _compute_severity(likelihood_value, impact_value)

            ok, assigned_to_norm = _validate_assignee_email(assigned_to)
            if not ok:
                allowlist, _src = _allowed_assignee_emails()
                if assigned_to_norm and allowlist:
                    flash("Assignee must be a member of the Risk Management group.", "warning")
                else:
                    flash("Assignee must be a valid email address (e.g. name@hdh.org).", "warning")
                return redirect(url_for("admin_risk_detail", risk_id=risk_id))
            assigned_to = assigned_to_norm
            close_reason = (close_reason or "").strip()

            if status.strip() == "Closed" and not close_reason:
                flash("Close reason is required when closing a risk.", "warning")
                return redirect(url_for("admin_risk_detail", risk_id=risk_id))

            actor = _current_actor()

            now = datetime.utcnow().isoformat()

            def _as_str(value) -> str:
                return ("" if value is None else str(value)).strip()

            updates_to_track = {
                "status": (risk["status"], status),
                "severity": (risk["severity"], severity),
                "assigned_to": (risk["assigned_to"], assigned_to),
                "owner": (risk["owner"], owner),
                "priority": (risk["priority"], priority),
                "progress": (risk["progress"], max(0, min(progress_value, 100))),
                "close_reason": (risk["close_reason"], close_reason),
            }
            for field, (old_val, new_val) in updates_to_track.items():
                if _as_str(old_val) != _as_str(new_val):
                    _log_event(
                        conn,
                        risk_id=risk_id,
                        event_type="update",
                        field=field,
                        old_value=_as_str(old_val),
                        new_value=_as_str(new_val),
                        actor=actor,
                    )

            # Notifications: assignment change
            old_assignee = _normalize_email(_as_str(risk["assigned_to"]))
            new_assignee = _normalize_email(_as_str(assigned_to))
            if new_assignee and new_assignee != old_assignee:
                try:
                    if _env_bool("NOTIFY_ASSIGNMENT_ENABLED", True):
                        base_url = os.getenv("APP_BASE_URL", "").strip() or request.url_root.rstrip("/")
                        link = f"{base_url}{url_for('admin_risk_detail', risk_id=risk_id)}"
                        owner_email = _normalize_email(owner)
                        to_addrs = _unique_emails([new_assignee, owner_email])
                        sent = False
                        if to_addrs:
                            sent = _send_email(
                                subject=f"Risk Assigned: {risk['title']}",
                                html_body=(
                                    f"<p>You have been assigned a risk.</p>"
                                    f"<p><strong>Title:</strong> {risk['title']}</p>"
                                    f"<p><strong>Status:</strong> {status}</p>"
                                    f"<p><strong>Severity:</strong> {severity or 'Unspecified'}</p>"
                                    f"<p><a href=\"{link}\">Open risk</a></p>"
                                ),
                                to_addrs=to_addrs,
                            )
                        if sent:
                            _log_event(
                                conn,
                                risk_id=risk_id,
                                event_type="notify",
                                field="email",
                                old_value="",
                                new_value=",".join(to_addrs)[:500],
                                actor=actor,
                            )
                except Exception:
                    pass

            # Auto-stamp lifecycle dates
            closed_at = (risk["closed_at"] or "").strip()
            if status.strip() == "Closed" and not closed_at:
                closed_at = now
            if status.strip() != "Closed":
                # Keep history but don't force closed_at for non-closed statuses
                pass

            date_mitigated_value = (date_mitigated or "").strip()
            if status.strip() == "Mitigated" and not date_mitigated_value:
                date_mitigated_value = now[:10]

            conn.execute(
                """
                UPDATE risks
                SET status = ?, owner = ?, mitigation = ?, admin_notes = ?,
                    progress = ?, priority = ?, impact_type = ?,
                    likelihood = ?, impact = ?,
                    likelihood_initial = ?, impact_initial = ?,
                    likelihood_residual = ?, impact_residual = ?,
                    date_identified = ?, review_period = ?,
                    date_last_reviewed = ?, next_review = ?,
                    date_postponed = ?, reason_postponed = ?, date_mitigated = ?,
                    mitigation_history = ?,
                    severity = ?, assigned_to = ?, updated_by = ?,
                    closed_at = ?, close_reason = ?,
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    status,
                    owner,
                    mitigation,
                    admin_notes,
                    max(0, min(progress_value, 100)),
                    priority,
                    impact_type,
                    likelihood_value,
                    impact_value,
                    likelihood_initial,
                    impact_initial,
                    likelihood_residual,
                    impact_residual,
                    date_identified,
                    review_period,
                    date_last_reviewed,
                    next_review,
                    date_postponed,
                    reason_postponed,
                    date_mitigated_value,
                    mitigation_history,
                    severity,
                    assigned_to,
                    actor,
                    closed_at,
                    close_reason,
                    now,
                    risk_id,
                ),
            )
            conn.commit()
            flash("Risk updated.", "success")
            return redirect(url_for("admin_risk_detail", risk_id=risk_id))

    return render_template(
        "admin_risk_detail.html",
        risk=risk,
        events=events,
        events_pagination=events_pagination,
        attachments=attachments,
        tasks=tasks,
        watchers=watchers,
        comments=comments,
        severity_options=SEVERITY_OPTIONS,
        assignees=sorted(set(assignees)),
        owners=sorted(set(owners)),
        task_status_options=TASK_STATUS_OPTIONS,
    )


@app.route("/admin/risk/<int:risk_id>/delete", methods=["POST"])
def admin_risk_delete(risk_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    reason = (request.form.get("reason") or "").strip()
    if len(reason) > 500:
        reason = reason[:500]

    actor = _current_actor()
    now = datetime.utcnow().isoformat()

    with get_db_connection() as conn:
        risk = conn.execute("SELECT * FROM risks WHERE id = ?", (risk_id,)).fetchone()
        if risk is None:
            flash("Risk not found.", "warning")
            return redirect(url_for("admin_dashboard"))

        if (risk["deleted_at"] or "").strip():
            flash("Risk is already deleted.", "info")
            return redirect(url_for("admin_risk_detail", risk_id=risk_id))

        conn.execute(
            """
            UPDATE risks
            SET deleted_at = ?, deleted_by = ?, deleted_reason = ?,
                updated_by = ?, updated_at = ?
            WHERE id = ?
            """.strip(),
            (now, actor, reason, actor, now, risk_id),
        )

        new_value = now if not reason else f"{now} | {reason}"[:500]
        _log_event(
            conn,
            risk_id=risk_id,
            event_type="delete",
            field="deleted_at",
            old_value="",
            new_value=new_value,
            actor=actor,
        )
        conn.commit()

    flash("Risk deleted (soft delete).", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/risk/<int:risk_id>/restore", methods=["POST"])
def admin_risk_restore(risk_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    actor = _current_actor()
    now = datetime.utcnow().isoformat()

    with get_db_connection() as conn:
        risk = conn.execute("SELECT * FROM risks WHERE id = ?", (risk_id,)).fetchone()
        if risk is None:
            flash("Risk not found.", "warning")
            return redirect(url_for("admin_dashboard"))

        old_deleted_at = (risk["deleted_at"] or "").strip()
        if not old_deleted_at:
            flash("Risk is not deleted.", "info")
            return redirect(url_for("admin_risk_detail", risk_id=risk_id))

        conn.execute(
            """
            UPDATE risks
            SET deleted_at = NULL, deleted_by = NULL, deleted_reason = NULL,
                updated_by = ?, updated_at = ?
            WHERE id = ?
            """.strip(),
            (actor, now, risk_id),
        )

        _log_event(
            conn,
            risk_id=risk_id,
            event_type="restore",
            field="deleted_at",
            old_value=old_deleted_at,
            new_value="",
            actor=actor,
        )
        conn.commit()

    flash("Risk restored.", "success")
    return redirect(url_for("admin_risk_detail", risk_id=risk_id))


@app.route("/admin/risk/<int:risk_id>/attachments/upload", methods=["POST"])
def admin_upload_attachments(risk_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    files = request.files.getlist("files")
    if not files:
        flash("No files selected.", "warning")
        return redirect(url_for("admin_risk_detail", risk_id=risk_id))

    actor = _current_actor()
    now = datetime.utcnow().isoformat()
    risk_dir = _risk_upload_path(risk_id)
    risk_dir.mkdir(parents=True, exist_ok=True)

    saved = 0
    with get_db_connection() as conn:
        exists = conn.execute(
            "SELECT 1 FROM risks WHERE id = ?",
            (risk_id,),
        ).fetchone()
        if not exists:
            flash("Risk not found.", "warning")
            return redirect(url_for("admin_dashboard"))

        for f in files:
            if not f or not getattr(f, "filename", ""):
                continue
            original = secure_filename(f.filename)
            if not original:
                continue
            if not _allowed_upload(original):
                flash(f"File type not allowed: {original}", "warning")
                continue

            ext = Path(original).suffix.lower()
            storage_name = f"{uuid.uuid4().hex}{ext}"
            path = risk_dir / storage_name
            f.save(path)
            size = path.stat().st_size if path.exists() else None

            conn.execute(
                """
                INSERT INTO risk_attachments (risk_id, storage_name, original_name, content_type, size_bytes, uploaded_by, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    risk_id,
                    storage_name,
                    original,
                    getattr(f, "mimetype", ""),
                    size,
                    actor,
                    now,
                ),
            )
            _log_event(
                conn,
                risk_id=risk_id,
                event_type="attachment",
                field="attachment",
                old_value="",
                new_value=original,
                actor=actor,
            )
            saved += 1

        conn.execute(
            "UPDATE risks SET updated_by = ?, updated_at = ? WHERE id = ?",
            (actor, now, risk_id),
        )
        conn.commit()

    flash(f"Uploaded {saved} file(s).", "success" if saved else "warning")
    return redirect(url_for("admin_risk_detail", risk_id=risk_id))


@app.route("/admin/risk/<int:risk_id>/attachments/<int:attachment_id>/delete", methods=["POST"])
def admin_delete_attachment(risk_id: int, attachment_id: int):
    if not require_admin():
        return redirect(url_for("admin"))

    actor = _current_actor()
    now = datetime.utcnow().isoformat()

    with get_db_connection() as conn:
        attachment = conn.execute(
            "SELECT * FROM risk_attachments WHERE id = ? AND risk_id = ?",
            (attachment_id, risk_id),
        ).fetchone()
        if attachment is None:
            flash("Attachment not found.", "warning")
            return redirect(url_for("admin_risk_detail", risk_id=risk_id))

        storage_name = attachment["storage_name"]
        original_name = attachment["original_name"] or storage_name
        is_archived = int(attachment["is_archived"] or 0) == 1
        base_dir = _risk_archive_path(risk_id) if is_archived else _risk_upload_path(risk_id)
        file_path = base_dir / storage_name
        deleted_file = False
        try:
            if file_path.exists() and file_path.is_file():
                file_path.unlink()
                deleted_file = True
        except Exception:
            # Keep going: DB record deletion is still useful.
            deleted_file = False

        conn.execute(
            "DELETE FROM risk_attachments WHERE id = ? AND risk_id = ?",
            (attachment_id, risk_id),
        )
        _log_event(
            conn,
            risk_id=risk_id,
            event_type="attachment_delete",
            field="attachment",
            old_value=original_name,
            new_value="",
            actor=actor,
        )
        conn.execute(
            "UPDATE risks SET updated_by = ?, updated_at = ? WHERE id = ?",
            (actor, now, risk_id),
        )
        conn.commit()

    flash(
        f"Deleted attachment: {original_name}" + ("" if deleted_file else " (file missing on disk)"),
        "success",
    )
    return redirect(url_for("admin_risk_detail", risk_id=risk_id))


if __name__ == "__main__":
    init_db()
    seed_from_excel()
    app.run(
        host=os.getenv("FLASK_HOST", "0.0.0.0"),
        port=int(os.getenv("FLASK_PORT", "5000")),
        debug=os.getenv("FLASK_DEBUG", "true").lower() == "true",
    )


